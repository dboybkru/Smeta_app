
import json
import base64
import hashlib
import hmac
import os
import re
import time
from datetime import datetime, timezone
from html import escape
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker

from crud import (
    add_smeta_item,
    create_smeta_revision,
    clone_smeta,
    create_material,
    create_smeta,
    delete_smeta,
    delete_smeta_item,
    filter_materials,
    get_material,
    get_materials,
    get_smeta,
    get_smeta_revisions,
    get_smetas,
    get_visible_smetas,
    restore_smeta_revision,
    normalize_search_text,
    update_smeta,
    update_smeta_item,
)
from models import Base, Smeta, SmetaAccess, SmetaRevision, User


def simple_ai_assistant(prompt: str):
    text = prompt.lower()
    if "дорого" in text or "цена" in text:
        return "Рассмотрите более дешевые материалы или альтернативные поставщики."
    if "ошибка" in text or "провер" in text:
        return "Проверьте единицы измерения, дубли позиций и строки с нулевым количеством."
    return "Рассмотрите оптимизацию сметы: проверить единицы и объемы."


def append_ai_audit(event_type, payload):
    record = {
        "ts": datetime.now(timezone.utc).isoformat(),
        "event": event_type,
        **payload,
    }
    try:
        AI_AUDIT_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with AI_AUDIT_LOG_PATH.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
    except Exception:
        pass


def log_ai_command(stage, user, prompt, payload=None, reply=None, results=None, selected_smeta_id=None, extra=None):
    append_ai_audit(
        "ai_command",
        {
            "stage": stage,
            "user": getattr(user, "email", None),
            "user_id": getattr(user, "id", None),
            "prompt": prompt,
            "payload": payload,
            "reply": reply,
            "results": results or [],
            "selected_smeta_id": selected_smeta_id,
            "extra": extra or {},
        },
    )


DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./smeta.db")
SETTINGS_PATH = Path(__file__).with_name("settings.json")
AI_AUDIT_LOG_PATH = Path(__file__).with_name("logs") / "assistant_audit.jsonl"
AUTH_SECRET = os.getenv("AUTH_SECRET", "local-smeta-secret-change-me")
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "dboy@bk.ru").lower()
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "avigYUHv1")
DEFAULT_SECTIONS = [
    "Оборудование",
    "Монтажные работы",
    "Пусконаладочные работы",
    "Кабельные линии",
    "Материалы и расходники",
    "Доставка и логистика",
    "Проектирование",
    "Прочее",
]
DEVICE_SYNONYMS = {
    "камера": ["камера", "камеры", "камеру", "камерой", "видеокамера", "видеокамеры", "видеокамеру"],
    "регистратор": [
        "регистратор",
        "регистратора",
        "регистраторы",
        "видеорегистратор",
        "видеорегистратора",
        "nvr",
        "dvr",
    ],
    "коммутатор": ["коммутатор", "коммутатора", "switch"],
    "турникет": ["турникет", "турникета"],
    "считыватель": ["считыватель", "считывателя"],
    "кнопка выхода": ["кнопка", "кнопки", "кнопка выхода", "выход"],
    "датчик": ["датчик", "датчика", "извещатель", "извещателя"],
    "блок питания": [
        "блок питания",
        "источник питания",
        "источник бесперебойного питания",
        "источника питания",
        "источника бесперебойного питания",
        "ибп",
    ],
    "видеосервер": ["видеосервер", "видеосервера", "сервер"],
    "жесткий диск": ["жд", "hdd", "диск", "жесткий диск", "жёсткий диск"],
    "автоматика ворот": ["ворота", "воротами", "автоматика ворот", "комплект управления воротами"],
    "контроллер скуд": ["контроллер", "скуд", "с2000-2"],
    "преобразователь": ["преобразователь", "rs-485", "rs-232", "ethernet"],
    "замок": ["замок", "электромагнитный замок"],
    "смк": ["смк", "сигнализатор магнитоконтактный"],
    "кронштейн": ["кронштейн", "кронштейна"],
    "кабель": ["кабель", "кабеля"],
    "короб": ["короб", "короба"],
    "гофротруба": ["гофротруба", "гофротрубы", "гофра"],
    "аренда вышки": ["аренда вышки", "вышка"],
}
SYSTEM_DEFINITIONS = {
    "Видеонаблюдение": {"камера", "регистратор", "видеосервер"},
    "СКУД": {
        "турникет",
        "считыватель",
        "кнопка выхода",
        "контроллер скуд",
        "преобразователь",
        "замок",
        "смк",
        "автоматика ворот",
    },
}
COMMISSIONING_TEMPLATE_NAMES = {
    "Видеонаблюдение": "Пусконаладка системы видеонаблюдения",
    "СКУД": "Пусконаладка системы СКУД",
}
WORK_TEMPLATE_NAMES = {
    "камера": "Монтаж камеры",
    "регистратор": "Монтаж видеорегистратора",
    "коммутатор": "Монтаж коммутатора",
    "турникет": "Монтаж турникета",
    "считыватель": "Монтаж считывателя",
    "кнопка выхода": "Монтаж кнопки выхода",
    "датчик": "Монтаж датчика",
    "блок питания": "Монтаж блока питания",
    "видеосервер": "Монтаж видеосервера",
    "жесткий диск": "Монтаж жесткого диска",
    "автоматика ворот": "Монтаж автоматики ворот",
    "контроллер скуд": "Монтаж контроллера СКУД",
    "преобразователь": "Монтаж преобразователя интерфейсов",
    "замок": "Монтаж электромагнитного замка",
    "смк": "Монтаж СМК",
    "кронштейн": "Монтаж кронштейна",
    "кабель": "Прокладка кабеля",
    "короб": "Монтаж короба",
    "гофротруба": "Прокладка гофротрубы",
    "аренда вышки": "Аренда вышки",
}
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base.metadata.create_all(bind=engine)


def ensure_schema():
    with engine.begin() as conn:
        tables = {
            "materials": ["characteristics", "item_type"],
            "works": ["characteristics"],
            "smeta_items": ["characteristics", "section", "base_unit_price"],
            "smetas": [
                "parent_id",
                "owner_id",
                "customer_name",
                "customer_details",
                "contractor_name",
                "contractor_details",
                "approver_name",
                "approver_details",
                "tax_mode",
                "tax_rate",
                "section_adjustments",
            ],
        }
        for table_name, columns in tables.items():
            existing = {row[1] for row in conn.execute(text(f"PRAGMA table_info({table_name})"))}
            for column in columns:
                if column not in existing:
                    column_type = "FLOAT" if column in {"base_unit_price"} else "VARCHAR"
                    conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column} {column_type}"))
        conn.execute(
            text(
                "UPDATE materials SET item_type = CASE "
                "WHEN lower(coalesce(source, '')) LIKE '%работ%' "
                "OR lower(coalesce(name, '')) LIKE 'монтаж %' "
                "OR coalesce(name, '') LIKE 'Монтаж %' "
                "OR lower(coalesce(name, '')) LIKE 'демонтаж %' "
                "OR coalesce(name, '') LIKE 'Демонтаж %' "
                "OR lower(coalesce(name, '')) LIKE 'прокладка %' "
                "OR coalesce(name, '') LIKE 'Прокладка %' "
                "OR lower(coalesce(name, '')) LIKE 'установка %' "
                "OR coalesce(name, '') LIKE 'Установка %' "
                "OR lower(coalesce(name, '')) LIKE 'настройка %' "
                "OR coalesce(name, '') LIKE 'Настройка %' "
                "OR lower(coalesce(name, '')) LIKE 'подключение %' "
                "OR coalesce(name, '') LIKE 'Подключение %' "
                "OR coalesce(name, '') LIKE 'Аренда вышки%' "
                "THEN 'work' ELSE 'equipment' END"
            )
        )
        conn.execute(
            text(
                "UPDATE smeta_items SET base_unit_price = unit_price "
                "WHERE base_unit_price IS NULL"
            )
        )


ensure_schema()


def b64url(data):
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode("ascii")


def b64url_decode(data):
    padding = "=" * (-len(data) % 4)
    return base64.urlsafe_b64decode((data + padding).encode("ascii"))


def hash_password(password, salt=None):
    salt_bytes = salt or os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt_bytes, 200000)
    return f"pbkdf2${b64url(salt_bytes)}${b64url(digest)}"


def verify_password(password, stored_hash):
    try:
        _, salt_value, digest_value = (stored_hash or "").split("$", 2)
        expected = b64url_decode(digest_value)
        actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), b64url_decode(salt_value), 200000)
        return hmac.compare_digest(actual, expected)
    except (ValueError, TypeError):
        return False


def create_token(user):
    payload = {"sub": user.id, "email": user.email, "is_admin": bool(user.is_admin), "exp": int(time.time()) + 60 * 60 * 24 * 14}
    body = b64url(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
    signature = b64url(hmac.new(AUTH_SECRET.encode("utf-8"), body.encode("ascii"), hashlib.sha256).digest())
    return f"{body}.{signature}"


def decode_token(token):
    try:
        body, signature = token.split(".", 1)
        expected = b64url(hmac.new(AUTH_SECRET.encode("utf-8"), body.encode("ascii"), hashlib.sha256).digest())
        if not hmac.compare_digest(signature, expected):
            return None
        payload = json.loads(b64url_decode(body))
        if int(payload.get("exp", 0)) < int(time.time()):
            return None
        return payload
    except (ValueError, TypeError, json.JSONDecodeError):
        return None


def ensure_admin_user():
    db = SessionLocal()
    try:
        admin = db.query(User).filter(User.email == ADMIN_EMAIL).first()
        if not admin:
            admin = User(email=ADMIN_EMAIL, password_hash=hash_password(ADMIN_PASSWORD), is_admin=1)
            db.add(admin)
            db.commit()
            db.refresh(admin)
        elif not verify_password(ADMIN_PASSWORD, admin.password_hash):
            admin.password_hash = hash_password(ADMIN_PASSWORD)
            admin.is_admin = 1
            db.commit()
            db.refresh(admin)
        db.execute(text("UPDATE smetas SET owner_id = :owner_id WHERE owner_id IS NULL"), {"owner_id": admin.id})
        db.commit()
    finally:
        db.close()


ensure_admin_user()


def ensure_revision_seed():
    db = SessionLocal()
    try:
        smetas = db.query(Smeta).order_by(Smeta.id.asc()).all()
        for smeta in smetas:
            existing = db.query(SmetaRevision).filter(SmetaRevision.smeta_id == smeta.id).first()
            if not existing:
                create_smeta_revision(db, smeta, "seed")
    finally:
        db.close()


ensure_revision_seed()

app = FastAPI(title="Сметное приложение с AI")

allowed_origins = os.getenv("CORS_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[origin.strip() for origin in allowed_origins.split(",") if origin.strip()],
    allow_methods=["*"],
    allow_headers=["*"],
)


class MaterialIn(BaseModel):
    item_type: str = "equipment"
    name: str = Field(..., min_length=1)
    characteristics: str = ""
    unit: str = ""
    price: float = Field(..., ge=0)
    source: str = ""


class SmetaIn(BaseModel):
    parent_id: int | None = None
    name: str = Field(..., min_length=1)
    customer_name: str = ""
    customer_details: str = ""
    contractor_name: str = ""
    contractor_details: str = ""
    approver_name: str = ""
    approver_details: str = ""
    tax_mode: str = "none"
    tax_rate: float = Field(0, ge=0, le=100)
    section_adjustments: dict[str, float] = {}


class AuthIn(BaseModel):
    email: str = Field(..., min_length=3)
    password: str = Field(..., min_length=1)


class RegisterIn(BaseModel):
    email: str = Field(..., min_length=3)
    password: str = Field(..., min_length=6)


class ShareIn(BaseModel):
    email: str = Field(..., min_length=3)
    permission: str = "view"


class SmetaItemIn(BaseModel):
    item_type: str = "material"
    section: str = "Оборудование"
    name: str = Field(..., min_length=1)
    characteristics: str = ""
    unit: str = ""
    quantity: float = Field(1, gt=0)
    unit_price: float = Field(..., ge=0)
    source: str = ""


class AiSettingsIn(BaseModel):
    base_url: str = "https://api.vsegpt.ru/v1"
    api_key: str = ""
    model: str = ""
    assistant_prompt: str = ""


class AiCommandIn(BaseModel):
    prompt: str = Field(..., min_length=1)
    smeta_id: int | None = None


class PriceImportResult(BaseModel):
    status: str
    imported: int
    skipped: int = 0


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def user_to_dict(user):
    created_at = getattr(user, "created_at", None)
    return {
        "id": user.id,
        "email": user.email,
        "is_admin": bool(user.is_admin),
        "created_at": created_at.isoformat() if created_at else None,
    }


def require_admin_user(user):
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail="Только для администратора")
    return user


def get_current_user(authorization: str = Header(default=""), db: Session = Depends(get_db)):
    scheme, _, token = (authorization or "").partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=401, detail="Нужна авторизация")
    payload = decode_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Сессия истекла")
    user = db.query(User).filter(User.id == int(payload.get("sub") or 0)).first()
    if not user:
        raise HTTPException(status_code=401, detail="Пользователь не найден")
    return user


def smeta_permission(db, smeta_id, user):
    smeta = get_smeta(db, smeta_id)
    if not smeta:
        return None, ""
    if user.is_admin or normalized_owner_id(smeta) == user.id:
        return smeta, "owner"
    access = (
        db.query(SmetaAccess)
        .filter(SmetaAccess.smeta_id == smeta_id, SmetaAccess.user_id == user.id)
        .first()
    )
    return smeta, (access.permission if access else "")


def require_smeta_access(db, smeta_id, user, write=False):
    smeta, permission = smeta_permission(db, smeta_id, user)
    if not smeta:
        raise HTTPException(status_code=404, detail="Смета не найдена")
    if permission == "owner" or (permission == "edit") or (permission == "view" and not write):
        return smeta
    raise HTTPException(status_code=403, detail="Нет доступа к смете")


def material_to_dict(material):
    item_type = material.item_type or classify_catalog_item(material.name, material.source)
    return {
        "id": material.id,
        "item_type": item_type,
        "name": material.name,
        "characteristics": material.characteristics or "",
        "unit": material.unit or "",
        "price": material.price,
        "source": material.source or "",
        "last_update": material.last_update,
    }


def parse_section_adjustments(value):
    if isinstance(value, dict):
        raw = value
    else:
        try:
            raw = json.loads(value or "{}")
        except (TypeError, json.JSONDecodeError):
            raw = {}
    adjustments = {}
    for section, percent in raw.items():
        try:
            adjustments[str(section)] = max(-100.0, min(1000.0, float(percent or 0)))
        except (TypeError, ValueError):
            adjustments[str(section)] = 0
    return adjustments


def section_adjustment_percent(smeta, section):
    return parse_section_adjustments(getattr(smeta, "section_adjustments", "{}")).get(section or "Оборудование", 0)


def effective_unit_price(item, smeta):
    percent = section_adjustment_percent(smeta, item.section or "Оборудование")
    return round((item.unit_price or 0) * (1 + percent / 100), 2)


def item_total(item, smeta):
    return round((item.quantity or 0) * effective_unit_price(item, smeta), 2)


def smeta_financials(smeta):
    subtotal = round(sum(item_total(item, smeta) for item in smeta.items), 2)
    tax_mode = getattr(smeta, "tax_mode", "none") or "none"
    tax_rate = float(getattr(smeta, "tax_rate", 0) or 0)
    if tax_mode == "vat_added" and tax_rate > 0:
        tax_amount = round(subtotal * tax_rate / 100, 2)
        total = round(subtotal + tax_amount, 2)
    elif tax_mode == "vat_included" and tax_rate > 0:
        total = subtotal
        tax_amount = round(subtotal * tax_rate / (100 + tax_rate), 2)
    else:
        tax_amount = 0
        total = subtotal
    return {"subtotal": subtotal, "tax_amount": tax_amount, "total": total}


def normalized_parent_id(smeta):
    try:
        return int(getattr(smeta, "parent_id", None) or 0) or None
    except (TypeError, ValueError):
        return None


def normalized_owner_id(smeta):
    try:
        return int(getattr(smeta, "owner_id", None) or 0) or None
    except (TypeError, ValueError):
        return None


def item_to_dict(item, smeta=None):
    price = effective_unit_price(item, smeta) if smeta else (item.unit_price or 0)
    base_value = getattr(item, "base_unit_price", None)
    base_price = float(base_value if base_value is not None else (item.unit_price or 0))
    return {
        "id": item.id,
        "item_type": item.item_type,
        "section": item.section or "Оборудование",
        "name": item.name,
        "characteristics": item.characteristics or "",
        "unit": item.unit or "",
        "quantity": item.quantity,
        "unit_price": item.unit_price,
        "base_unit_price": base_price,
        "effective_unit_price": price,
        "section_adjustment_percent": section_adjustment_percent(smeta, item.section or "Оборудование") if smeta else 0,
        "source": item.source or "",
        "total": round((item.quantity or 0) * price, 2),
    }


def smeta_to_dict(smeta):
    items = [item_to_dict(item, smeta) for item in smeta.items]
    financials = smeta_financials(smeta)
    return {
        "id": smeta.id,
        "parent_id": normalized_parent_id(smeta),
        "owner_id": normalized_owner_id(smeta),
        "is_branch": bool(normalized_parent_id(smeta)),
        "name": smeta.name,
        "customer_name": smeta.customer_name or "",
        "customer_details": smeta.customer_details or "",
        "contractor_name": smeta.contractor_name or "",
        "contractor_details": smeta.contractor_details or "",
        "approver_name": smeta.approver_name or "",
        "approver_details": smeta.approver_details or "",
        "tax_mode": getattr(smeta, "tax_mode", "none") or "none",
        "tax_rate": float(getattr(smeta, "tax_rate", 0) or 0),
        "section_adjustments": parse_section_adjustments(getattr(smeta, "section_adjustments", "{}")),
        "created_at": smeta.created_at,
        "items": items,
        "subtotal": financials["subtotal"],
        "tax_amount": financials["tax_amount"],
        "total": financials["total"],
    }


def http_error_detail(exc, fallback):
    response = getattr(exc, "response", None)
    if response is None:
        return f"{fallback}: {exc}"
    try:
        data = response.json()
        detail = data.get("error", data.get("detail", data))
        if isinstance(detail, dict):
            detail = detail.get("message", json.dumps(detail, ensure_ascii=False))
    except (ValueError, AttributeError):
        detail = response.text
    return f"{fallback}: HTTP {response.status_code}. {str(detail)[:600]}"


def read_settings():
    defaults = {
        "base_url": "https://api.vsegpt.ru/v1",
        "api_key": "",
        "model": "",
        "assistant_prompt": (
            "Ты встроенный ассистент сметчика. Отвечай кратко, по-русски и по делу. "
            "Работай как практик по сметам: смотри на смету целиком, а не на один раздел, "
            "связывай оборудование, монтаж и пусконаладку по смыслу устройства и системы. "
            "Если видишь регистратор, видеорегистратор, NVR или DVR, предлагай монтаж и пусконаладку "
            "системы видеонаблюдения, а при необходимости настройку удаленного доступа. "
            "Если видишь СКУД-оборудование, связывай его с монтажом и пусконаладкой СКУД. "
            "Не придумывай цены, не дублируй позиции без необходимости и не теряй количество. "
            "Если данных не хватает, задай один короткий уточняющий вопрос. "
            "Если пользователь просит проверить, исправить или пересчитать смету, проверь всю смету целиком."
        ),
    }
    if not SETTINGS_PATH.exists():
        return defaults
    try:
        data = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return defaults
    return {**defaults, **data}


def write_settings(settings):
    SETTINGS_PATH.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def merged_ai_system_prompt(settings):
    prompt = (settings.get("assistant_prompt") or "").strip()
    return prompt or (
        "Ты встроенный ассистент сметчика. Отвечай кратко, по-русски и по делу. "
        "Если пользователь просит проверить смету, смотри на нее целиком. "
        "Не придумывай цены и не дублируй позиции без необходимости."
    )


def public_settings(settings):
    api_key = settings.get("api_key", "")
    masked_key = f"{api_key[:6]}...{api_key[-4:]}" if len(api_key) > 12 else bool(api_key)
    return {
        "base_url": settings.get("base_url", ""),
        "model": settings.get("model", ""),
        "has_api_key": bool(api_key),
        "masked_api_key": masked_key,
        "assistant_prompt": settings.get("assistant_prompt", ""),
    }


def provider_headers(settings):
    if not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="Сначала сохраните API-ключ в настройках")
    return {"Authorization": f"Bearer {settings['api_key']}", "Content-Type": "application/json"}


def strip_code_fences(text_value):
    text_value = (text_value or "").strip()
    fence_match = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", text_value, flags=re.IGNORECASE)
    if fence_match:
        return fence_match.group(1).strip()
    return text_value


def extract_balanced_json_object(text_value):
    text_value = text_value or ""
    start = text_value.find("{")
    if start < 0:
        return None
    depth = 0
    in_string = False
    escaped = False
    for index in range(start, len(text_value)):
        char = text_value[index]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
        else:
            if char == '"':
                in_string = True
            elif char == "{":
                depth += 1
            elif char == "}":
                depth -= 1
                if depth == 0:
                    return text_value[start : index + 1]
    return None


def escape_json_control_chars(text_value):
    text_value = text_value or ""
    result = []
    in_string = False
    escaped = False
    for char in text_value:
        if in_string:
            if escaped:
                result.append(char)
                escaped = False
                continue
            if char == "\\":
                result.append(char)
                escaped = True
                continue
            if char == '"':
                result.append(char)
                in_string = False
                continue
            if char == "\n":
                result.append("\\n")
                continue
            if char == "\r":
                result.append("\\r")
                continue
            if char == "\t":
                result.append("\\t")
                continue
        else:
            if char == '"':
                in_string = True
        result.append(char)
    return "".join(result)


def parse_ai_object(content):
    if isinstance(content, dict):
        content.setdefault("reply", "")
        content.setdefault("actions", [])
        return content
    if not isinstance(content, str):
        return None
    candidates = []
    stripped = strip_code_fences(content)
    candidates.append(stripped)
    balanced = extract_balanced_json_object(stripped)
    if balanced and balanced not in candidates:
        candidates.append(balanced)
    for candidate in candidates:
        for variant in (candidate, escape_json_control_chars(candidate)):
            try:
                data = json.loads(variant)
            except json.JSONDecodeError:
                continue
            if isinstance(data, dict):
                data.setdefault("reply", "")
                data.setdefault("actions", [])
                return data
    return None


def endpoint(base_url, path):
    return f"{base_url.rstrip('/')}/{path.lstrip('/')}"


def extract_price(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        numbers = re.findall(r"\d+(?:[.,]\d+)?", value)
        return float(numbers[0].replace(",", ".")) if numbers else None
    return None


def extract_strict_price(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    if isinstance(value, str):
        text_value = value.strip().replace("\u00a0", " ")
        text_value = re.sub(r"\s+", "", text_value)
        text_value = text_value.replace("₽", "").replace("руб.", "").replace("руб", "")
        text_value = text_value.replace(",", ".")
        if re.fullmatch(r"\d+(?:\.\d+)?", text_value):
            number = float(text_value)
            return number if number > 0 else None
    return None


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def classify_catalog_item(name, source=""):
    text_value = f"{name} {source}".lower()
    name_value = (name or "").lower().strip()
    work_starts = (
        "монтаж ",
        "демонтаж ",
        "прокладка ",
        "установка ",
        "настройка ",
        "подключение ",
        "пусконаладка ",
        "пусконаладочные ",
        "обслуживание ",
    )
    if "работ" in (source or "").lower() or name_value.startswith(work_starts):
        return "work"
    return "equipment"


def default_section_for_type(item_type):
    return "Монтажные работы" if item_type == "work" else "Оборудование"


def summarize_characteristics(text_value, max_lines=3, line_length=95):
    parts = [part.strip(" .;") for part in re.split(r"[.;]\s+", text_value or "") if part.strip()]
    if not parts:
        return ""
    lines = []
    for part in parts:
        if len(part) > line_length:
            part = part[: line_length - 1].rstrip() + "…"
        lines.append(part)
        if len(lines) >= max_lines:
            break
    return "\n".join(lines)


def normalize_quantity(value):
    try:
        quantity = int(round(float(value or 1)))
    except (TypeError, ValueError):
        quantity = 1
    return max(quantity, 1)


def is_numeric_price(value):
    price = extract_strict_price(value)
    return price is not None and price > 0


def normalize_label(value):
    return clean_text(value).lower().replace("\n", " ")


def get_nested(data, paths):
    for path in paths:
        value = data
        for part in path:
            if not isinstance(value, dict) or part not in value:
                value = None
                break
            value = value[part]
        if value is not None:
            return value
    return None


def normalize_model(model):
    input_price = get_nested(
        model,
        [
            ("pricing", "prompt"),
            ("pricing", "input"),
            ("price", "prompt"),
            ("metadata", "prompt_price"),
            ("metadata", "input_price"),
        ],
    )
    output_price = get_nested(
        model,
        [
            ("pricing", "completion"),
            ("pricing", "output"),
            ("price", "completion"),
            ("metadata", "completion_price"),
            ("metadata", "output_price"),
        ],
    )
    return {
        "id": model.get("id") or model.get("name"),
        "name": model.get("name") or model.get("id"),
        "input_price": extract_price(input_price),
        "output_price": extract_price(output_price),
        "raw_pricing": model.get("pricing") or model.get("price") or model.get("metadata", {}),
    }


def compact_text(text, limit=12000):
    text = re.sub(r"\n{3,}", "\n\n", text)
    if len(text) <= limit:
        return text
    return text[:limit] + "\n\n[Текст обрезан приложением из-за лимита модели]"


def call_ai_json(system_prompt, user_text):
    settings = read_settings()
    if not settings.get("model"):
        raise HTTPException(status_code=400, detail="Выберите модель AI в настройках")
    payload = {
        "model": settings["model"],
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": compact_text(user_text)},
        ],
    }
    try:
        with httpx.Client(timeout=90) as client:
            response = client.post(
                endpoint(settings["base_url"], "chat/completions"),
                headers=provider_headers(settings),
                json=payload,
            )
            response.raise_for_status()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=http_error_detail(exc, "AI-провайдер отклонил запрос")) from exc

    content = response.json()["choices"][0]["message"]["content"]
    match = re.search(r"\[[\s\S]*\]", content)
    if not match:
        raise HTTPException(status_code=502, detail="AI не вернул JSON-массив материалов")
    try:
        return json.loads(match.group(0))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail="AI вернул невалидный JSON") from exc


def call_ai_object(system_prompt, user_text):
    settings = read_settings()
    if not settings.get("model"):
        raise HTTPException(status_code=400, detail="Выберите модель AI в настройках")
    payload = {
        "model": settings["model"],
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": compact_text(user_text, 16000)},
        ],
    }
    try:
        with httpx.Client(timeout=90) as client:
            response = client.post(
                endpoint(settings["base_url"], "chat/completions"),
                headers=provider_headers(settings),
                json=payload,
            )
            response.raise_for_status()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=http_error_detail(exc, "AI-провайдер отклонил запрос")) from exc

    content = response.json()["choices"][0]["message"]["content"]
    parsed = parse_ai_object(content)
    if parsed:
        return parsed
    return {"reply": content.strip(), "actions": []}


class TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []

    def handle_data(self, data):
        text_part = data.strip()
        if text_part:
            self.parts.append(text_part)

    def text(self):
        return "\n".join(self.parts)


def html_to_text(html):
    parser = TextExtractor()
    parser.feed(html)
    return parser.text()


def dataframe_to_text(df):
    df = df.dropna(how="all").fillna("")
    return df.head(160).to_csv(index=False)


def first_matching_column(columns, variants):
    normalized = [(column, str(column).lower()) for column in columns]
    for column, lower_column in normalized:
        if any(variant in lower_column for variant in variants):
            return column
    return None


def header_score(row):
    labels = [normalize_label(value) for value in row if clean_text(value)]
    text_row = " | ".join(labels)
    score = 0
    if any(word in text_row for word in ["наименование", "название", "номенклатура", "товар", "работ"]):
        score += 3
    if any(word in text_row for word in ["цена", "цены", "розн", "опт", "парт", "стоим"]):
        score += 3
    if any(word in text_row for word in ["ед.", "ед ", "изм", "руб./", "вал."]):
        score += 1
    if any(word in text_row for word in ["описание", "характерист", "краткие"]):
        score += 1
    return score


def find_header_row(rows):
    best_index = None
    best_score = 0
    for index, row in enumerate(rows[:80]):
        score = header_score(row)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index if best_score >= 4 else None


def column_values(rows, column_index, start_index, limit=80):
    values = []
    for row in rows[start_index : start_index + limit]:
        if column_index < len(row):
            values.append(row[column_index])
    return values


def detect_name_column(headers, rows, data_start):
    candidates = []
    for index, header in enumerate(headers):
        label = normalize_label(header)
        score = 0
        if any(word in label for word in ["наименование", "название", "номенклатура", "товар"]):
            score += 100
        if "работ" in label:
            score += 80
        values = column_values(rows, index, data_start)
        text_count = sum(1 for value in values if len(clean_text(value)) >= 3 and not is_numeric_price(value))
        score += text_count
        if score:
            candidates.append((score, index))
    return max(candidates)[1] if candidates else None


def detect_characteristics_column(headers):
    for index, header in enumerate(headers):
        label = normalize_label(header)
        if any(word in label for word in ["характер", "описан", "кратк", "параметр", "модель"]):
            return index
    return None


def detect_unit_column(headers):
    for index, header in enumerate(headers):
        label = normalize_label(header)
        if any(word in label for word in ["ед.", "ед ", "изм", "вал./", "руб./"]):
            return index
    return None


def price_header_priority(header):
    label = normalize_label(header)
    if "парт" in label:
        return 100
    if "кр.опт" in label or "круп" in label:
        return 95
    if "опт" in label:
        return 90
    if "цена" in label or "цены" in label or "стоим" in label:
        return 75
    if "инст" in label:
        return 55
    if "розн" in label:
        return 35
    return 0


def detect_price_columns(headers, parent_headers, rows, data_start):
    candidates = []
    for index, header in enumerate(headers):
        combined_header = f"{clean_text(parent_headers[index])} {clean_text(header)}"
        header_priority = price_header_priority(combined_header)
        values = column_values(rows, index, data_start, limit=120)
        numeric_count = sum(1 for value in values if is_numeric_price(value))
        if header_priority or numeric_count >= 3:
            if any(word in normalize_label(combined_header) for word in ["код", "артикул", "№", "номер"]):
                continue
            score = header_priority + numeric_count
            candidates.append((score, index))
    return [index for _, index in sorted(candidates, reverse=True)]


def likely_category_row(row, name_index, price_indexes):
    name = clean_text(row[name_index] if name_index is not None and name_index < len(row) else "")
    if not name:
        return False
    has_price = any(index < len(row) and is_numeric_price(row[index]) for index in price_indexes)
    filled = sum(1 for value in row if clean_text(value))
    return not has_price and filled <= 3 and len(name) >= 3


def pick_price(row, price_indexes):
    prices = []
    for index in price_indexes:
        if index < len(row):
            price = extract_strict_price(row[index])
            if price is not None and price > 0:
                prices.append(price)
    if not prices:
        return None
    return min(prices)


def parse_excel_workbook(file_obj, source):
    file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    parsed = []
    for sheet in workbook.worksheets:
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        header_index = find_header_row(rows)
        if header_index is None:
            continue

        parent_headers = rows[header_index]
        next_headers = rows[header_index + 1] if header_index + 1 < len(rows) else ()
        use_next_for_price = any(normalize_label(value) == "цены" for value in parent_headers)
        headers = []
        for index, header in enumerate(parent_headers):
            next_header = next_headers[index] if index < len(next_headers) else ""
            headers.append(f"{clean_text(header)} {clean_text(next_header) if use_next_for_price else ''}".strip())

        data_start = header_index + (2 if use_next_for_price else 1)
        name_index = detect_name_column(headers, rows, data_start)
        if name_index is None:
            continue
        characteristics_index = detect_characteristics_column(headers)
        unit_index = detect_unit_column(headers)
        price_indexes = detect_price_columns(headers, parent_headers, rows, data_start)
        if not price_indexes:
            continue

        current_group = sheet.title
        for row in rows[data_start:]:
            row = tuple(row)
            if likely_category_row(row, name_index, price_indexes):
                current_group = clean_text(row[name_index])
                continue

            name = clean_text(row[name_index] if name_index < len(row) else "")
            price = pick_price(row, price_indexes)
            if not name or price is None:
                continue
            if name.lower() in {"наименование", "название", "товар", "цена"}:
                continue

            characteristics = (
                clean_text(row[characteristics_index])
                if characteristics_index is not None and characteristics_index < len(row)
                else ""
            )
            unit = clean_text(row[unit_index]) if unit_index is not None and unit_index < len(row) else ""
            if unit.lower().startswith("руб./"):
                unit = unit.split("/", 1)[-1]
            parsed.append(
                {
                    "name": name,
                    "characteristics": characteristics or current_group,
                    "unit": unit,
                    "price": price,
                    "source": f"{source} / {sheet.title}",
                    "item_type": classify_catalog_item(name, f"{source} / {sheet.title}"),
                }
            )
    return parsed


def save_parsed_materials(db, rows):
    imported = 0
    skipped = 0
    seen = set()
    for row in rows:
        key = (row["name"], row.get("source", ""), row.get("price"))
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        create_material(
            db,
            row["name"],
            row.get("unit", ""),
            float(row["price"]),
            row.get("source", ""),
            row.get("characteristics", ""),
            row.get("item_type", "equipment"),
        )
        imported += 1
    return imported, skipped


def import_excel_by_guess(db, df, source):
    name_column = first_matching_column(
        df.columns,
        ["name", "назван", "наимен", "номенклат", "товар", "материал", "позиция"],
    )
    price_column = first_matching_column(
        df.columns,
        ["price", "цена", "стоим", "прайс", "розн", "опт"],
    )
    unit_column = first_matching_column(df.columns, ["unit", "ед", "изм"])
    characteristics_column = first_matching_column(
        df.columns,
        ["характер", "описан", "параметр", "модель", "артикул"],
    )

    if not name_column or not price_column:
        return 0, 0

    imported = 0
    skipped = 0
    for _, row in df.dropna(how="all").iterrows():
        name = "" if pd.isna(row.get(name_column)) else str(row.get(name_column)).strip()
        price = extract_price(row.get(price_column))
        if not name or price is None:
            skipped += 1
            continue
        unit = "" if not unit_column or pd.isna(row.get(unit_column)) else str(row.get(unit_column)).strip()
        characteristics = (
            ""
            if not characteristics_column or pd.isna(row.get(characteristics_column))
            else str(row.get(characteristics_column)).strip()
        )
        create_material(db, name, unit, price, source, characteristics, classify_catalog_item(name, source))
        imported += 1
    return imported, skipped


async def extract_pdf_text(file):
    settings = read_settings()
    if not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="Для PDF нужен API-ключ AI-провайдера")
    content = await file.read()
    try:
        async with httpx.AsyncClient(timeout=90) as client:
            response = await client.post(
                endpoint(settings["base_url"], "extract_text"),
                headers={"Authorization": f"Bearer {settings['api_key']}"},
                files={"file": (file.filename, content, file.content_type or "application/pdf")},
            )
            response.raise_for_status()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=http_error_detail(exc, "Не удалось извлечь текст из PDF")) from exc
    data = response.json()
    if isinstance(data, dict):
        return data.get("text") or data.get("content") or data.get("result") or json.dumps(data, ensure_ascii=False)
    return str(data)


def save_ai_materials(db, rows, source):
    imported = 0
    skipped = 0
    for row in rows:
        if not isinstance(row, dict):
            skipped += 1
            continue
        name = str(row.get("name") or row.get("название") or "").strip()
        price = extract_price(row.get("price") or row.get("цена"))
        if not name or price is None:
            skipped += 1
            continue
        create_material(
            db,
            name,
            str(row.get("unit") or row.get("единица") or "").strip(),
            price,
            str(row.get("source") or source or "").strip(),
            str(row.get("characteristics") or row.get("характеристики") or "").strip(),
            classify_catalog_item(name, source),
        )
        imported += 1
    return imported, skipped


def smeta_context(db):
    smetas = get_smetas(db)
    return [
        {
            "id": smeta.id,
            "name": smeta.name,
            "total": smeta_to_dict(smeta)["total"],
            "items": [
                {
                    "id": item.id,
                    "item_type": item.item_type,
                    "section": item.section or "Оборудование",
                    "name": item.name,
                    "quantity": item.quantity,
                    "unit": item.unit or "",
                    "unit_price": item.unit_price,
                }
                for item in smeta.items
            ],
        }
        for smeta in smetas[:20]
    ]


def smeta_equipment_summary(smeta):
    summary = {}
    for item in smeta.items:
        if (item.item_type or "") == "work" or (item.section or "") in {"Монтажные работы", "Пусконаладочные работы"}:
            continue
        kind = device_kind(item.name)
        if not kind:
            continue
        if kind not in summary:
            summary[kind] = {"quantity": 0, "items": []}
        summary[kind]["quantity"] += item.quantity or 0
        summary[kind]["items"].append(
            {
                "id": item.id,
                "name": item.name,
                "quantity": item.quantity or 0,
                "section": item.section or "",
            }
        )
    return summary


def is_equipment_smeta_item(item):
    if (item.item_type or "") == "work":
        return False
    if (item.section or "") in {"Монтажные работы", "Пусконаладочные работы"}:
        return False
    return bool(device_kind(item.name))


def answer_count_question(prompt, smeta):
    if not smeta:
        return None
    text_value = (prompt or "").lower()
    if not any(word in text_value for word in ["сколько", "количество", "скока"]):
        return None
    requested_kind = None
    for kind, terms in DEVICE_SYNONYMS.items():
        if any(term in text_value for term in terms):
            requested_kind = kind
            break
    if not requested_kind:
        return None
    summary = smeta_equipment_summary(smeta)
    data = summary.get(requested_kind, {"quantity": 0, "items": []})
    item_text = "; ".join(f"{item['name']} x{item['quantity']:g}" for item in data["items"])
    synonym_note = ""
    if requested_kind == "камера":
        synonym_note = " Камера и видеокамера считаются одним типом."
    elif requested_kind == "регистратор":
        synonym_note = " Регистратор и видеорегистратор считаются одним типом."
    return f"{requested_kind.capitalize()}: {data['quantity']:g} шт.{synonym_note}" + (
        f" Позиции: {item_text}." if item_text else ""
    )


def tokenize(text_value):
    return [
        token
        for token in re.findall(r"[a-zа-яё0-9]+", (text_value or "").lower())
        if len(token) >= 3
    ]


def device_kind(name):
    text_value = (name or "").lower()
    for kind, words in DEVICE_SYNONYMS.items():
        if any(word in text_value for word in words):
            return kind
    tokens = tokenize(name)
    return tokens[0] if tokens else ""


def find_work_price(db, equipment_name):
    kind = device_kind(equipment_name)
    search_terms = DEVICE_SYNONYMS.get(kind, [kind]) if kind else []
    candidate_by_id = {}
    for term in search_terms:
        for work in get_materials(db, term, "work", 80):
            candidate_by_id[work.id] = work
    candidates = list(candidate_by_id.values())
    install_words = ["монтаж", "установка", "прокладка", "подключение"]

    def score(work):
        name = (work.name or "").lower()
        value = 0
        if any(term in name for term in search_terms):
            value += 50
        if any(word in name for word in install_words):
            value += 30
        if name.startswith("монтаж"):
            value += 20
        if "демонтаж" in name:
            value -= 60
        return value

    candidates = sorted(candidates, key=lambda work: (-score(work), work.price or 0))
    if candidates and score(candidates[0]) > 0:
        return candidates[0], kind
    return None, kind


def work_item_data_for_equipment(db, equipment_name, equipment_quantity=1, equipment_names=None):
    work, kind = find_work_price(db, equipment_name)
    quantity = normalize_quantity(equipment_quantity)
    related_names = equipment_names or equipment_name
    if work:
        return {
            "item_type": "work",
            "section": "Монтажные работы",
            "name": work.name,
            "characteristics": f"Для: {related_names}",
            "unit": work.unit or "шт",
            "quantity": quantity,
            "unit_price": work.price,
            "source": work.source or "База работ",
        }, kind, True
    fallback_name = WORK_TEMPLATE_NAMES.get(kind, f"Монтаж {kind or equipment_name}")
    return {
        "item_type": "work",
        "section": "Монтажные работы",
        "name": fallback_name,
        "characteristics": f"Цена не найдена. Для: {related_names}",
        "unit": "шт",
        "quantity": quantity,
        "unit_price": 0,
        "source": "Нет цены в базе работ",
    }, kind, False


def remember_work_price_from_smeta(db, work_item):
    if not work_item or (work_item.unit_price or 0) <= 0:
        return None
    kind = work_kind(work_item)
    name = WORK_TEMPLATE_NAMES.get(kind, work_item.name or f"Монтаж {kind}")
    if not name:
        return None
    existing, _ = find_work_price(db, name)
    if existing:
        existing.price = float(work_item.unit_price or 0)
        existing.unit = work_item.unit or existing.unit or "шт"
        existing.source = work_item.source or existing.source or "Из сметы"
        if work_item.characteristics and not existing.characteristics:
            existing.characteristics = work_item.characteristics
        db.commit()
        db.refresh(existing)
        return existing
    return create_material(
        db,
        name,
        work_item.unit or "шт",
        float(work_item.unit_price or 0),
        work_item.source or "Из сметы",
        work_item.characteristics or "",
        "work",
    )


def is_commissioning_item(item):
    text_value = f"{item.name or ''} {item.characteristics or ''}".lower()
    return (item.section or "") == "Пусконаладочные работы" or "пусконалад" in text_value


def detected_system_names(equipment_by_kind):
    systems = []
    kinds = set(equipment_by_kind.keys())
    for system_name, system_kinds in SYSTEM_DEFINITIONS.items():
        if kinds & system_kinds:
            systems.append(system_name)
    return systems


def commissioning_system_from_item(item):
    text_value = f"{item.name or ''} {item.characteristics or ''}".lower()
    if any(word in text_value for word in ["видеонаб", "камера", "видеокамера", "регистратор", "видеосервер"]):
        return "Видеонаблюдение"
    if any(word in text_value for word in ["скуд", "считыватель", "контроллер", "замок", "смк", "турникет"]):
        return "СКУД"
    return ""


def find_commissioning_price(db, system_name=None):
    candidates = []
    system_queries = []
    if system_name:
        system_queries = [system_name, COMMISSIONING_TEMPLATE_NAMES.get(system_name, "")]
    for query in ["пусконаладочные работы", "пусконаладка", "пнр", *system_queries]:
        candidates.extend(get_materials(db, query, "work", 20))
    system_terms = {
        "Видеонаблюдение": ["видеонаб", "камера", "видеокамера", "регистратор"],
        "СКУД": ["скуд", "считыватель", "контроллер", "замок"],
    }.get(system_name, [])

    def score(work):
        name = (work.name or "").lower()
        characteristics = (work.characteristics or "").lower()
        text_value = f"{name} {characteristics}"
        has_commissioning_word = "пусконалад" in text_value or "пнр" in text_value
        if not has_commissioning_word:
            return -100
        if system_terms and not any(term in text_value for term in system_terms):
            return -100
        value = 0
        if has_commissioning_word:
            value += 80
        if name.startswith("пусконалад"):
            value += 30
        if "работ" in text_value:
            value += 10
        if system_terms and any(term in text_value for term in system_terms):
            value += 60
        return value

    candidates = sorted({work.id: work for work in candidates}.values(), key=lambda work: (-score(work), work.price or 0))
    if candidates and score(candidates[0]) > 0:
        return candidates[0]
    return None


def find_work_price_by_query(db, query_text, required_terms=None):
    required_terms = required_terms or []
    candidates = get_materials(db, query_text, "work", 80)

    def score(work):
        name = (work.name or "").lower()
        characteristics = (work.characteristics or "").lower()
        text_value = f"{name} {characteristics}"
        value = 0
        if query_text and query_text.lower() in text_value:
            value += 80
        if any(term in text_value for term in required_terms):
            value += 70
        if name.startswith("настройка") or name.startswith("пусконалад"):
            value += 20
        if "доступ" in text_value:
            value += 25
        if "удален" in text_value:
            value += 25
        if "монтаж" in text_value:
            value -= 10
        return value

    candidates = sorted({work.id: work for work in candidates}.values(), key=lambda work: (-score(work), work.price or 0))
    if candidates and score(candidates[0]) > 0:
        return candidates[0]
    return None


def remember_commissioning_price_from_smeta(db, commissioning_item, system_name=None):
    if not commissioning_item or (commissioning_item.unit_price or 0) <= 0:
        return None
    existing = find_commissioning_price(db, system_name)
    name = COMMISSIONING_TEMPLATE_NAMES.get(system_name) or commissioning_item.name or "Пусконаладочные работы"
    if existing:
        existing.price = float(commissioning_item.unit_price or 0)
        existing.unit = commissioning_item.unit or existing.unit or "компл"
        existing.source = commissioning_item.source or existing.source or "Из сметы"
        if commissioning_item.characteristics and not existing.characteristics:
            existing.characteristics = commissioning_item.characteristics
        db.commit()
        db.refresh(existing)
        return existing
    return create_material(
        db,
        name,
        commissioning_item.unit or "компл",
        float(commissioning_item.unit_price or 0),
        commissioning_item.source or "Из сметы",
        commissioning_item.characteristics or "",
        "work",
    )


def ensure_commissioning_for_smeta(db, smeta_id, smeta, equipment_by_kind):
    systems = detected_system_names(equipment_by_kind)
    if not systems:
        systems = []

    results = []
    commissioning_items = [item for item in smeta.items if is_commissioning_item(item)]
    items_by_system = {}
    generic_items = []
    for item in commissioning_items:
        system_name = commissioning_system_from_item(item)
        if system_name:
            items_by_system.setdefault(system_name, []).append(item)
        else:
            generic_items.append(item)

    for system_name in systems:
        existing_items = items_by_system.get(system_name) or []
        primary = existing_items[0] if existing_items else (generic_items.pop(0) if generic_items else None)
        template = find_commissioning_price(db, system_name)
        primary_name = (primary.name or "").lower() if primary else ""
        primary_is_commissioning = "пусконалад" in primary_name or "пнр" in primary_name
        remembered = remember_commissioning_price_from_smeta(db, primary, system_name) if primary and primary_is_commissioning else None
        primary_name_has_system = any(
            term in primary_name
            for term in {
                "Видеонаблюдение": ["видеонаб", "камера", "видеокамера", "регистратор"],
                "СКУД": ["скуд", "считыватель", "контроллер", "замок"],
            }.get(system_name, [])
        )
        name = (
            primary.name
            if primary and commissioning_system_from_item(primary) and primary_is_commissioning and primary_name_has_system
            else (remembered.name if remembered else (template.name if template else COMMISSIONING_TEMPLATE_NAMES[system_name]))
        )
        unit = primary.unit if primary and primary.unit else (remembered.unit if remembered else (template.unit if template else "компл"))
        unit_price = (
            primary.unit_price
            if primary and primary_is_commissioning and (primary.unit_price or 0) > 0
            else (remembered.price if remembered else (template.price if template else 0))
        )
        source = (
            primary.source
            if primary and primary.source
            else (remembered.source if remembered else (template.source if template else "Нет цены в базе работ"))
        )
        item_data = {
            "item_type": "work",
            "section": "Пусконаладочные работы",
            "name": name,
            "characteristics": f"Система: {system_name}",
            "unit": unit,
            "quantity": 1,
            "unit_price": unit_price,
            "source": source,
        }
        if primary:
            update_smeta_item(
                db,
                smeta_id,
                primary.id,
                item_data,
            )
            results.append(f"Пусконаладка «{system_name}» уже есть, цена {unit_price}")
        else:
            add_smeta_item(db, smeta_id, item_data)
            results.append(f"Добавил пусконаладку «{system_name}» по {unit_price}")

        for duplicate in existing_items[1:]:
            delete_smeta_item(db, smeta_id, duplicate.id)
            results.append(f"Удалил дублирующую пусконаладку «{duplicate.name}»")

    for item in generic_items:
        if (item.unit_price or 0) <= 0:
            delete_smeta_item(db, smeta_id, item.id)
            results.append(f"Удалил лишнюю общую пусконаладку «{item.name}»")

    return results


def work_matches_equipment(work_item, equipment):
    kind = device_kind(equipment.name)
    if not kind:
        return False
    terms = DEVICE_SYNONYMS.get(kind, [kind])
    haystack = f"{work_item.name or ''} {work_item.characteristics or ''}".lower()
    if any(term in haystack for term in terms):
        return True
    return (equipment.name or "").lower() in haystack


def work_kind(item):
    text_value = f"{item.name or ''} {item.characteristics or ''}".lower()
    for kind, terms in DEVICE_SYNONYMS.items():
        if any(term in text_value for term in terms):
            return kind
    return device_kind(item.name)


def has_matching_work(items, equipment):
    equipment_kind = device_kind(equipment.name)
    for item in items:
        if (item.section or "") != "Монтажные работы":
            continue
        name = (item.name or "").lower()
        if work_kind(item) == equipment_kind:
            return True
        if work_matches_equipment(item, equipment) and any(
            word in name for word in ["монтаж", "установка", "прокладка", "подключение"]
        ):
            return True
    return False


def validate_and_fix_smeta(db, smeta_id):
    smeta = get_smeta(db, smeta_id)
    if not smeta:
        return None, ["Смета не найдена"]
    results = []
    equipment_items = [item for item in smeta.items if is_equipment_smeta_item(item)]
    work_items = [item for item in smeta.items if (item.section or "") == "Монтажные работы"]

    summary = smeta_equipment_summary(smeta)
    if summary:
        counts = ", ".join(f"{kind}: {data['quantity']:g}" for kind, data in summary.items())
        results.append(f"Оборудование по всей смете: {counts}")

    equipment_by_kind = {}
    for equipment in equipment_items:
        kind = device_kind(equipment.name)
        if not kind:
            continue
        bucket = equipment_by_kind.setdefault(kind, {"quantity": 0, "items": []})
        bucket["quantity"] += equipment.quantity or 0
        bucket["items"].append(equipment)

    for kind, data_by_kind in equipment_by_kind.items():
        quantity = data_by_kind["quantity"] or 1
        sample_equipment = data_by_kind["items"][0]
        item_data, _, found_price = work_item_data_for_equipment(
            db,
            sample_equipment.name,
            quantity,
            ", ".join(item.name for item in data_by_kind["items"]),
        )
        matching_items = [item for item in work_items if work_kind(item) == kind]

        if matching_items:
            primary = matching_items[0]
            if not found_price and (primary.unit_price or 0) > 0:
                remembered = remember_work_price_from_smeta(db, primary)
                item_data["name"] = remembered.name if remembered else item_data["name"]
                item_data["unit"] = remembered.unit if remembered else (primary.unit or item_data["unit"])
                item_data["unit_price"] = remembered.price if remembered else primary.unit_price
                item_data["source"] = remembered.source if remembered else (primary.source or "Ручная цена")
                if remembered:
                    results.append(f"Сохранил «{remembered.name}» в базе работ по цене {remembered.price}")
            update_data = {
                "item_type": "work",
                "section": "Монтажные работы",
                "name": item_data["name"],
                "characteristics": item_data["characteristics"],
                "unit": item_data["unit"],
                "quantity": item_data["quantity"],
                "unit_price": item_data["unit_price"],
                "source": item_data["source"],
            }
            update_smeta_item(db, smeta_id, primary.id, update_data)
            for duplicate in matching_items[1:]:
                delete_smeta_item(db, smeta_id, duplicate.id)
                results.append(f"Удалил дублирующий монтаж «{duplicate.name}»")
            price_text = update_data["unit_price"]
            results.append(f"Привёл монтаж типа «{kind}» к количеству {quantity:g} по цене {price_text}")
        else:
            if found_price:
                results.append(
                    f"Добавил «{item_data['name']}» для типа «{kind}» количеством {quantity:g} по {item_data['unit_price']}"
                )
            else:
                results.append(f"Добавил монтаж типа «{kind}» количеством {quantity:g} с ценой 0")
            add_smeta_item(db, smeta_id, item_data)

    smeta = get_smeta(db, smeta_id)
    results.extend(ensure_commissioning_for_smeta(db, smeta_id, smeta, equipment_by_kind))
    return get_smeta(db, smeta_id), results or ["Ошибок не найдено"]


def add_installation_works_for_smeta(db, smeta_id):
    smeta = get_smeta(db, smeta_id)
    if not smeta:
        return None, ["Смета не найдена"]
    results = []
    equipment_items = [item for item in smeta.items if is_equipment_smeta_item(item)]
    if not equipment_items:
        return smeta, ["В смете нет оборудования, для которого можно добавить монтаж"]

    for equipment in equipment_items:
        if has_matching_work(smeta.items, equipment):
            results.append(f"Монтаж для «{equipment.name}» уже есть")
            continue
        work, kind = find_work_price(db, equipment.name)
        quantity = equipment.quantity or 1
        if work:
            item_data = {
                "item_type": "work",
                "section": "Монтажные работы",
                "name": work.name,
                "characteristics": f"Для: {equipment.name}",
                "unit": work.unit or "шт",
                "quantity": quantity,
                "unit_price": work.price,
                "source": work.source or "База работ",
            }
            add_smeta_item(db, smeta_id, item_data)
            results.append(f"Добавил «{work.name}» для «{equipment.name}» по {work.price}")
        else:
            fallback_name = f"Монтаж {kind or equipment.name}"
            item_data = {
                "item_type": "work",
                "section": "Монтажные работы",
                "name": fallback_name,
                "characteristics": f"Цена не найдена. Для: {equipment.name}",
                "unit": equipment.unit or "шт",
                "quantity": quantity,
                "unit_price": 0,
                "source": "Нет цены в базе работ",
            }
            add_smeta_item(db, smeta_id, item_data)
            results.append(f"Добавил «{fallback_name}» с ценой 0")
    return get_smeta(db, smeta_id), results


def should_auto_add_installation(prompt):
    text_value = (prompt or "").lower()
    return (
        any(word in text_value for word in ["монтаж", "установлено", "установить", "установка"])
        and any(word in text_value for word in ["кажд", "устрой", "оборуд", "позици"])
    )


def should_validate_smeta(prompt):
    text_value = (prompt or "").lower()
    return "смет" in text_value and any(word in text_value for word in ["проверь", "исправ", "почин", "провер"])


def looks_like_new_smeta_request(prompt):
    text_value = normalize_search_text(prompt or "")
    if not text_value:
        return False
    negative_words = [
        "проверь",
        "проверить",
        "исправ",
        "почин",
        "удали",
        "откати",
        "посмотри",
        "посмотреть",
        "сравни",
        "сравнить",
    ]
    if any(word in text_value for word in negative_words):
        return False
    creation_words = [
        "создай",
        "создать",
        "создайте",
        "сделай",
        "сделать",
        "нужно",
        "надо",
        "требуется",
        "хочу",
        "установ",
        "смонтир",
        "подобрать",
        "рассчитать",
        "собери",
        "собрать",
        "спроектир",
    ]
    project_words = [
        "смет",
        "объект",
        "видеонаб",
        "камер",
        "видеокамер",
        "регистрат",
        "nvr",
        "dvr",
        "скуд",
        "считывател",
        "контроллер",
        "замок",
        "турникет",
        "ибп",
        "жд",
        "hdd",
        "жестк",
        "кабель",
        "монтаж",
        "пусконалад",
    ]
    has_creation = any(word in text_value for word in creation_words)
    has_project = any(word in text_value for word in project_words)
    return has_creation and has_project


def looks_like_extend_smeta_request(prompt):
    text_value = normalize_search_text(prompt or "")
    if not text_value:
        return False
    negative_words = [
        "проверь",
        "проверить",
        "исправ",
        "почин",
        "удали",
        "откати",
        "посмотри",
        "посмотреть",
        "сравни",
        "сравнить",
        "создай",
        "создать",
        "создайте",
    ]
    if any(word in text_value for word in negative_words):
        return False
    add_words = [
        "добавь",
        "добавить",
        "дополни",
        "дополнить",
        "ещё",
        "еще",
        "плюс",
        "впиши",
        "вставь",
        "увеличь",
        "увеличить",
        "докинь",
    ]
    project_words = [
        "камер",
        "видеокамер",
        "регистрат",
        "видеонаб",
        "nvr",
        "dvr",
        "скуд",
        "считывател",
        "контроллер",
        "замок",
        "турникет",
        "ибп",
        "жд",
        "hdd",
        "кабель",
        "монтаж",
        "пусконалад",
        "коммутатор",
    ]
    return any(word in text_value for word in add_words) and any(word in text_value for word in project_words)


def auto_smeta_prefix(user):
    local_part = str(getattr(user, "email", "") or "dboy").split("@", 1)[0] or "dboy"
    safe = re.sub(r"[^0-9a-zA-Zа-яА-Я]+", "_", local_part).strip("_").lower()
    return safe or "dboy"


def next_auto_smeta_name(db, user):
    prefix = auto_smeta_prefix(user)
    pattern = f"{prefix}_auto_%"
    existing_names = [
        name[0]
        for name in db.query(Smeta.name).filter(Smeta.name.like(pattern)).all()
        if name and name[0]
    ]
    highest = 0
    prefix_token = f"{prefix}_auto_"
    for existing in existing_names:
        suffix = existing[len(prefix_token) :]
        match = re.match(r"(\d+)", suffix)
        if match:
            highest = max(highest, int(match.group(1)))
    return f"{prefix}_auto_{highest + 1}"


def prompt_requests_named_smeta(prompt):
    text_value = normalize_search_text(prompt or "")
    if not text_value:
        return False
    explicit_markers = [
        "назови",
        "назвать",
        "название",
        "под названием",
        "с названием",
        "имя сметы",
        "как назвать",
        "именуй",
    ]
    return any(marker in text_value for marker in explicit_markers)


def resolve_ai_smeta_name(db, prompt, raw_name, user):
    fallback_name = str(getattr(user, "email", "") or "Новая смета").split("@", 1)[0] or "Новая смета"
    candidate = sanitize_smeta_name(raw_name or "", prompt, user)
    if prompt_requests_named_smeta(prompt) and candidate and candidate != fallback_name:
        return candidate
    return next_auto_smeta_name(db, user)


def smeta_has_ip_cameras(db, smeta_id):
    smeta = get_smeta(db, smeta_id)
    if not smeta:
        return False
    for item in getattr(smeta, "items", []) or []:
        if item.item_type and item.item_type != "equipment":
            continue
        name_text = normalize_search_text(item.name or "")
        characteristics_text = normalize_search_text(item.characteristics or "")
        if "ip" in name_text and ("камер" in name_text or "видеокамер" in name_text):
            return True
        if "ip" in characteristics_text and ("камер" in characteristics_text or "видеокамер" in characteristics_text):
            return True
    return False


def auto_build_project_smeta(db, smeta_id, prompt, user=None):
    text_value = (prompt or "").lower()
    results = []
    requests = []
    ip_context = "ip" in text_value or "ип" in text_value or smeta_has_ip_cameras(db, smeta_id)

    def add_request(query, quantity, label):
        requests.append({"query": query, "quantity": quantity, "label": label})

    def quantity_before(pattern, default=1):
        match = re.search(rf"(?:(\d+)\s*[xх]\s*)?(?:{pattern})", text_value)
        if match and match.group(1):
            return normalize_quantity(match.group(1))
        match = re.search(rf"(?:{pattern})\s*(\d+)\s*[xх]?", text_value)
        if match and match.group(1):
            return normalize_quantity(match.group(1))
        return default

    camera_match = re.search(r"(?:(\d+)\s*)?(?:ip|ип)?\s*(?:камер[аы]?|видеокамер[аы]?)", text_value)
    if camera_match:
        camera_qty = normalize_quantity(camera_match.group(1) or 1)
        camera_query = "ip камера"
        mp_match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:mp|мп|vg)", text_value)
        if mp_match:
            camera_query += f" {mp_match.group(1).replace(',', '.')} мп"
        add_request(camera_query, camera_qty, "камер")

    recorder_match = re.search(r"(?:(\d+)\s*)?(?:видеорегистратор|регистратор|nvr|dvr)", text_value)
    if recorder_match:
        recorder_qty = normalize_quantity(recorder_match.group(1) or 1)
        recorder_query = "ip видеорегистратор" if ip_context else "видеорегистратор"
        if "poe" in text_value or "пое" in text_value:
            recorder_query += " poe"
        add_request(recorder_query, recorder_qty, "видеорегистратора")

    hdd_match = re.search(r"(?:жд|hdd|жестк[а-я ]*диск[а-я ]*)(?:\s*(\d+(?:[.,]\d+)?))?\s*(?:тб|tb)", text_value)
    if hdd_match or any(term in text_value for term in ["жд", "hdd", "жесткий диск", "жёсткий диск"]):
        size_match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:тб|tb)", text_value)
        hdd_query = "жесткий диск"
        if size_match:
            hdd_query += f" {size_match.group(1).replace(',', '.')} тб"
        add_request(hdd_query, 1, "жесткого диска")

    cable_match = re.search(r"(?:кабель|кабел[ья])\s*(\d+(?:[.,]\d+)?)?\s*(?:м|метр[а-я]*)?", text_value)
    cable_qty_match = cable_match.group(1) if cable_match else None
    if not cable_qty_match:
        cable_qty_match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:м|метр[а-я]*)\s*(?:кабель|кабел[ья])", text_value)
        cable_qty_match = cable_qty_match.group(1) if cable_qty_match else None
    if cable_match or any(term in text_value for term in ["кабель", "кабеля", "кабелем"]):
        cable_qty = normalize_quantity(cable_qty_match or 1)
        cable_query = "кабель"
        if "витая пара" in text_value or "utp" in text_value:
            cable_query = "кабель витая пара"
        elif "силов" in text_value:
            cable_query = "кабель силовой"
        elif "гофр" in text_value:
            cable_query = "гофротруба"
        add_request(cable_query, cable_qty, "кабеля")

    ups_match = re.search(
        r"(?:(\d+)\s*)?(?:ибп|ups|источник(?:а|ов)?\s+бесперебойного\s+питания|источн[а-я ]*питани[яе]|блок питания)",
        text_value,
    )
    if ups_match or any(term in text_value for term in ["ибп", "ups", "источник бесперебойного питания", "блок питания"]):
        ups_qty = normalize_quantity(ups_match.group(1) or 1) if ups_match else 1
        add_request("ибп", ups_qty, "ИБП")

    switch_match = re.search(r"(?:коммутатор|switch)(?:\s*(\d+)\s*порто?)?", text_value)
    if switch_match or any(term in text_value for term in ["коммутатор", "switch"]):
        switch_qty = normalize_quantity(1)
        switch_query = "коммутатор"
        ports_match = re.search(r"(\d+)\s*порто", text_value)
        if ports_match:
            switch_query += f" {ports_match.group(1)} порт"
        if "poe" in text_value or "пое" in text_value:
            switch_query += " poe"
        add_request(switch_query, switch_qty, "коммутатора")

    controller_terms = r"(?:c2000-2|с2000-2|контроллер(?:а|ов|ы)?(?:\s+скуд)?|скуд)"
    if re.search(controller_terms, text_value):
        controller_qty = quantity_before(controller_terms, 1)
        add_request("c2000-2", controller_qty, "контроллера СКУД")

    reader_terms = r"(?:считывател[ьяеи]?|ридер[аы]?|reader[а-я ]*)"
    if re.search(reader_terms, text_value):
        reader_qty = quantity_before(reader_terms, 1)
        add_request("считыватель", reader_qty, "считывателя")

    lock_terms = r"(?:замок[а-я ]*|электромагнитн[а-я ]*замок[а-я ]*)"
    if re.search(lock_terms, text_value):
        lock_qty = quantity_before(lock_terms, 1)
        add_request("замок", lock_qty, "замка")

    exit_button_terms = r"(?:кнопк[а-я ]*выход[а-я ]*|выход)"
    if re.search(exit_button_terms, text_value) and "кнопк" in text_value:
        button_qty = quantity_before(exit_button_terms, 1)
        add_request("кнопка выхода", button_qty, "кнопки выхода")

    seen = set()
    for request in requests:
        key = (request["query"], request["quantity"])
        if key in seen:
            continue
        seen.add(key)
        materials = get_materials(db, request["query"], "equipment", 10)
        material = materials[0] if materials else None
        if material:
            item_data = {
                "item_type": material.item_type or "equipment",
                "section": default_section_for_type(material.item_type or "equipment"),
                "name": material.name,
                "characteristics": summarize_characteristics(material.characteristics or ""),
                "unit": material.unit or "",
                "quantity": request["quantity"],
                "unit_price": material.price,
                "source": material.source or "База материалов",
            }
            add_smeta_item(db, smeta_id, item_data)
            results.append(f"Добавил {request['label']} «{material.name}» x{request['quantity']}")
        else:
            item_data = {
                "item_type": "equipment",
                "section": "Оборудование",
                "name": request["query"],
                "characteristics": "Цена не найдена в базе.",
                "unit": "шт",
                "quantity": request["quantity"],
                "unit_price": 0,
                "source": "Нет цены в базе",
            }
            add_smeta_item(db, smeta_id, item_data)
            results.append(f"Добавил {request['label']} «{request['query']}» x{request['quantity']} с ценой 0")

    if requests:
        validate_and_fix_smeta(db, smeta_id)
    return results


def should_create_smeta(prompt):
    return looks_like_new_smeta_request(prompt)


def infer_smeta_name(prompt, reply, user=None):
    text_sources = [prompt or ""]
    fallback_name = str(getattr(user, "email", "") or "Новая смета").split("@", 1)[0] or "Новая смета"
    patterns = [
        r"смет[ауеы]?\s*(?:создан[аоы]?|готов[аоы]?|назов[иите]?|под названием)\s*[«\"']?([^\\n\\r\"'«»]+)",
        r"создай(?:те)?\s+смет[ауеы]?\s*(?:на|для)?\s*[«\"']?([^\\n\\r\"'«»]+)",
        r"смет[ауеы]?\s*(?:на|для)\s*[«\"']?([^\\n\\r\"'«»]+)",
    ]
    project_words = [
        "камер",
        "видеокамер",
        "регистрат",
        "видеонаб",
        "nvr",
        "dvr",
        "скуд",
        "считывател",
        "контроллер",
        "замок",
        "турникет",
        "ибп",
        "жд",
        "hdd",
        "кабель",
        "монтаж",
        "пусконалад",
        "коммутатор",
    ]

    def looks_like_title(candidate):
        normalized = normalize_search_text(candidate)
        if not normalized:
            return False
        if len(candidate.split()) > 8:
            return False
        if candidate.count(",") >= 1:
            return False
        if any(char.isdigit() for char in candidate) and any(word in normalized for word in project_words):
            return False
        if any(word in normalized for word in project_words):
            return False
        return True

    for text in text_sources:
        for pattern in patterns:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                name = re.sub(r"\s+", " ", match.group(1)).strip(" .,:;\"'«»")
                if name and name.lower() not in {"без названия", "без имени", "новая", "новую", "новая смета", "смета"} and looks_like_title(name):
                    return name[:80]
    return fallback_name


def sanitize_smeta_name(name, prompt=None, user=None):
    candidate = re.sub(r"\s+", " ", str(name or "")).strip(" .,:;\"'«»")
    if candidate:
        normalized = normalize_search_text(candidate)
        if len(candidate.split()) <= 8 and candidate.count(",") == 0:
            if not any(char.isdigit() for char in candidate):
                if not any(word in normalized for word in [
                    "камер",
                    "видеокамер",
                    "регистрат",
                    "видеонаб",
                    "nvr",
                    "dvr",
                    "скуд",
                    "считывател",
                    "контроллер",
                    "замок",
                    "турникет",
                    "ибп",
                    "жд",
                    "hdd",
                    "кабель",
                    "монтаж",
                    "пусконалад",
                    "коммутатор",
                ]):
                    return candidate[:80]
    inferred = infer_smeta_name(prompt or "", "", user)
    inferred = re.sub(r"\s+", " ", inferred).strip(" .,:;\"'«»")
    if inferred:
        return inferred[:80]
    return str(getattr(user, "email", "") or "Новая смета").split("@", 1)[0] or "Новая смета"


def execute_ai_actions(db, actions, fallback_smeta_id=None, user=None, prompt_text=""):
    results = []
    selected_smeta_id = fallback_smeta_id
    active_smeta_id = fallback_smeta_id
    for action in actions:
        if not isinstance(action, dict):
            continue
        action_type = action.get("action")
        if not action_type:
            continue
        if action_type == "create_smeta":
            name = resolve_ai_smeta_name(db, prompt_text, action.get("name"), user)
            smeta = create_smeta(db, name, {"owner_id": user.id} if user else None)
            selected_smeta_id = smeta.id
            active_smeta_id = smeta.id
            results.append(f"Создал смету «{name}»")
            if prompt_text and looks_like_new_smeta_request(prompt_text):
                build_results = auto_build_project_smeta(db, smeta.id, prompt_text, user)
                results.extend(build_results)
        elif action_type == "delete_smeta":
            smeta_id = int(action.get("smeta_id") or active_smeta_id or selected_smeta_id or 0)
            smeta = get_smeta(db, smeta_id)
            if user:
                require_smeta_access(db, smeta_id, user, write=True)
                if not user.is_admin and smeta and normalized_owner_id(smeta) != user.id:
                    results.append("Не удалил смету: удалять может только владелец")
                    continue
            if smeta and delete_smeta(db, smeta_id):
                results.append(f"Удалил смету «{smeta.name}»")
                if selected_smeta_id == smeta_id:
                    selected_smeta_id = None
            else:
                results.append("Не удалил смету: она не найдена")
        elif action_type == "add_item":
            smeta_id = int(action.get("smeta_id") or active_smeta_id or selected_smeta_id or 0)
            if active_smeta_id and not fallback_smeta_id:
                smeta_id = active_smeta_id
            if user:
                require_smeta_access(db, smeta_id, user, write=True)
            if not get_smeta(db, smeta_id):
                results.append("Не добавил позицию: смета не найдена")
                continue
            item_data = {
                "item_type": str(action.get("item_type") or "manual"),
                "section": str(action.get("section") or "Прочее"),
                "name": str(action.get("name") or "").strip(),
                "characteristics": str(action.get("characteristics") or "").strip(),
                "unit": str(action.get("unit") or "").strip(),
                "quantity": normalize_quantity(action.get("quantity") or 1),
                "unit_price": float(action.get("unit_price") or action.get("price") or 0),
                "source": str(action.get("source") or "AI").strip(),
            }
            if not item_data["name"]:
                results.append("Не добавил позицию: нет названия")
                continue
            add_smeta_item(db, smeta_id, item_data)
            results.append(f"Добавил позицию «{item_data['name']}»")
        elif action_type == "update_item":
            smeta_id = int(action.get("smeta_id") or active_smeta_id or selected_smeta_id or 0)
            if active_smeta_id and not fallback_smeta_id:
                smeta_id = active_smeta_id
            item_id = int(action.get("item_id") or 0)
            if user:
                require_smeta_access(db, smeta_id, user, write=True)
            data = {
                key: action.get(key)
                for key in ["item_type", "section", "name", "characteristics", "unit", "quantity", "unit_price", "source"]
                if key in action
            }
            if "quantity" in data:
                data["quantity"] = normalize_quantity(data["quantity"])
            if "unit_price" in data:
                data["unit_price"] = float(data["unit_price"])
            if update_smeta_item(db, smeta_id, item_id, data):
                results.append(f"Обновил позицию #{item_id}")
            else:
                results.append("Не обновил позицию: она не найдена")
        elif action_type == "delete_item":
            smeta_id = int(action.get("smeta_id") or active_smeta_id or selected_smeta_id or 0)
            if active_smeta_id and not fallback_smeta_id:
                smeta_id = active_smeta_id
            item_id = int(action.get("item_id") or 0)
            if user:
                require_smeta_access(db, smeta_id, user, write=True)
            if delete_smeta_item(db, smeta_id, item_id):
                results.append(f"Удалил позицию #{item_id}")
            else:
                results.append("Не удалил позицию: она не найдена")
        elif action_type not in {"noop", "none"}:
            results.append(f"Пропустил неизвестное действие: {action_type}")
    return selected_smeta_id, results


@app.get("/")
def read_root():
    return {"message": "Приложение для смет работает!"}


@app.post("/auth/login")
def login(payload: AuthIn, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == payload.email.strip().lower()).first()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Неверный логин или пароль")
    return {"access_token": create_token(user), "user": user_to_dict(user)}


@app.post("/auth/register")
def register(payload: RegisterIn, db: Session = Depends(get_db)):
    email = payload.email.strip().lower()
    if email == ADMIN_EMAIL:
        raise HTTPException(status_code=403, detail="Этот email зарезервирован для администратора")
    if db.query(User).filter(User.email == email).first():
        raise HTTPException(status_code=409, detail="Пользователь уже существует")
    user = User(email=email, password_hash=hash_password(payload.password), is_admin=0)
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"access_token": create_token(user), "user": user_to_dict(user)}


@app.delete("/admin/users/{user_id}")
def admin_delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_admin_user(user)
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if target.email == ADMIN_EMAIL:
        raise HTTPException(status_code=400, detail="Скрытого администратора удалить нельзя")
    if target.id == user.id:
        raise HTTPException(status_code=400, detail="Нельзя удалить самого себя")
    admin = db.query(User).filter(User.email == ADMIN_EMAIL).first()
    if admin:
        db.execute(text("UPDATE smetas SET owner_id = :owner_id WHERE owner_id = :target_id"), {"owner_id": admin.id, "target_id": target.id})
    db.query(SmetaAccess).filter(SmetaAccess.user_id == target.id).delete(synchronize_session=False)
    db.delete(target)
    db.commit()
    return {"status": "ok"}


@app.get("/auth/me")
def auth_me(user: User = Depends(get_current_user)):
    return user_to_dict(user)


@app.get("/admin/users")
def admin_users(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    require_admin_user(user)
    rows = db.query(User).order_by(User.id.desc()).all()
    return [user_to_dict(row) for row in rows]


@app.patch("/admin/users/{user_id}")
def admin_update_user(
    user_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_admin_user(user)
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if target.email == ADMIN_EMAIL and not bool(payload.get("is_admin", True)):
        raise HTTPException(status_code=400, detail="Скрытого администратора нельзя разжаловать")
    if "is_admin" in payload:
        target.is_admin = 1 if bool(payload.get("is_admin")) else 0
    db.commit()
    db.refresh(target)
    return user_to_dict(target)


@app.get("/admin/smetas/{smeta_id}/access")
def admin_smeta_access(
    smeta_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_admin_user(user)
    smeta = get_smeta(db, smeta_id)
    if not smeta:
        raise HTTPException(status_code=404, detail="Смета не найдена")
    rows = (
        db.query(SmetaAccess, User)
        .join(User, User.id == SmetaAccess.user_id)
        .filter(SmetaAccess.smeta_id == smeta_id)
        .order_by(User.email.asc())
        .all()
    )
    return {
        "smeta": smeta_to_dict(smeta),
        "access": [
            {
                "id": access.id,
                "user_id": user_row.id,
                "email": user_row.email,
                "permission": access.permission,
                "is_admin": bool(user_row.is_admin),
            }
            for access, user_row in rows
        ],
    }


@app.get("/admin/users/{user_id}/smetas")
def admin_user_smetas(
    user_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_admin_user(user)
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    access_rows = db.query(SmetaAccess).filter(SmetaAccess.user_id == user_id).all()
    access_map = {row.smeta_id: row.permission for row in access_rows}
    smetas = db.query(Smeta).order_by(Smeta.id.desc()).all()
    rows = []
    for smeta in smetas:
        owner_id = normalized_owner_id(smeta)
        permission = "owner" if owner_id == target.id else access_map.get(smeta.id)
        if not permission and not target.is_admin:
            continue
        if target.is_admin and not permission:
            permission = "admin"
        if permission:
            rows.append(
                {
                    "id": smeta.id,
                    "name": smeta.name,
                    "permission": permission,
                    "total": smeta_to_dict(smeta)["total"],
                    "parent_id": normalized_parent_id(smeta),
                }
            )
    return {"user": user_to_dict(target), "smetas": rows}


@app.delete("/admin/smetas/{smeta_id}/access/{user_id}")
def admin_revoke_access(
    smeta_id: int,
    user_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_admin_user(user)
    access = (
        db.query(SmetaAccess)
        .filter(SmetaAccess.smeta_id == smeta_id, SmetaAccess.user_id == user_id)
        .first()
    )
    if not access:
        raise HTTPException(status_code=404, detail="Доступ не найден")
    db.delete(access)
    db.commit()
    return {"status": "ok"}


@app.get("/sections")
def read_sections():
    return {"sections": DEFAULT_SECTIONS}


@app.get("/materials")
def read_materials(
    q: str = Query(default=""),
    item_type: str = Query(default="all"),
    technology: str = Query(default=""),
    megapixels: str = Query(default=""),
    price_to: float | None = Query(default=None, ge=0),
    limit: int = Query(default=200, ge=1, le=500),
    db: Session = Depends(get_db),
):
    rows = get_materials(db, q, item_type, limit=5000 if (technology or megapixels or price_to is not None) else limit)
    rows = filter_materials(rows, technology, megapixels, price_to)
    return [material_to_dict(material) for material in rows[:limit]]


@app.post("/materials")
def create_material_endpoint(material: MaterialIn, db: Session = Depends(get_db)):
    created = create_material(
        db,
        material.name.strip(),
        material.unit.strip(),
        material.price,
        material.source.strip(),
        material.characteristics.strip(),
        material.item_type,
    )
    return material_to_dict(created)


@app.post("/materials/import")
async def import_materials(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Загрузите Excel-файл .xlsx или .xls")

    try:
        rows = parse_excel_workbook(file.file, file.filename)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл") from exc

    imported, skipped = save_parsed_materials(db, rows)
    if imported == 0:
        raise HTTPException(
            status_code=422,
            detail="Не нашёл строки с названием и ценой. Проверьте, что в файле есть товарные строки и цены.",
        )
    return {"status": "ok", "imported": imported, "skipped": skipped}


@app.post("/materials/import-ai", response_model=PriceImportResult)
async def import_materials_with_ai(
    file: UploadFile | None = File(default=None),
    url: str = Form(default=""),
    db: Session = Depends(get_db),
):
    if not file and not url.strip():
        raise HTTPException(status_code=400, detail="Загрузите файл или укажите URL поставщика")

    source = url.strip() or (file.filename if file else "")
    text_content = ""
    excel_df = None

    if url.strip():
        try:
            async with httpx.AsyncClient(timeout=45, follow_redirects=True) as client:
                response = await client.get(url.strip())
                response.raise_for_status()
        except httpx.HTTPError as exc:
            raise HTTPException(status_code=400, detail=f"Не удалось загрузить сайт поставщика: {exc}") from exc
        text_content = html_to_text(response.text)
    elif file.filename.lower().endswith((".xlsx", ".xls")):
        try:
            parsed_rows = parse_excel_workbook(file.file, source)
            if parsed_rows:
                imported, skipped = save_parsed_materials(db, parsed_rows)
                return {"status": "ok", "imported": imported, "skipped": skipped}
            file.file.seek(0)
            excel_df = pd.read_excel(file.file)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл") from exc
        text_content = dataframe_to_text(excel_df)
    elif file.filename.lower().endswith(".pdf"):
        text_content = await extract_pdf_text(file)
    else:
        raise HTTPException(status_code=400, detail="Поддерживаются Excel, PDF или URL поставщика")

    if not text_content.strip():
        raise HTTPException(status_code=400, detail="Не удалось получить текст прайса")

    system_prompt = (
        "Ты извлекаешь строительные материалы из прайсов поставщиков. "
        "Верни только JSON-массив без markdown. Каждый объект строго с полями: "
        "name, characteristics, unit, price, source. "
        "price должен быть числом, unit короткой единицей измерения. "
        "Не выдумывай позиции и пропускай строки без цены."
    )
    rows = call_ai_json(system_prompt, f"Источник: {source}\n\nПрайс:\n{text_content}")
    imported, skipped = save_ai_materials(db, rows, source)
    if imported == 0 and excel_df is not None:
        imported, skipped = import_excel_by_guess(db, excel_df, source)
    if imported == 0:
        parsed_count = len(rows) if isinstance(rows, list) else 0
        raise HTTPException(
            status_code=422,
            detail=(
                "AI обработал прайс, но не нашёл материалы с названием и ценой. "
                f"JSON-строк от AI: {parsed_count}, пропущено: {skipped}. "
                "Попробуйте другую модель или Excel с колонками Наименование/Цена."
            ),
        )
    return {"status": "ok", "imported": imported, "skipped": skipped}


@app.get("/smetas")
def read_smetas(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return [smeta_to_dict(smeta) for smeta in get_visible_smetas(db, user)]


@app.post("/smetas")
def create_smeta_endpoint(payload: SmetaIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    data = payload.model_dump()
    name = data.pop("name").strip()
    data["tax_mode"] = data.get("tax_mode") if data.get("tax_mode") in {"none", "vat_added", "vat_included"} else "none"
    data["section_adjustments"] = json.dumps(
        parse_section_adjustments(data.get("section_adjustments", {})),
        ensure_ascii=False,
    )
    smeta = create_smeta(
        db,
        name,
        {
            **{key: (value if key in {"tax_rate", "parent_id"} else str(value).strip()) for key, value in data.items()},
            "owner_id": user.id,
        },
    )
    create_smeta_revision(db, get_smeta(db, smeta.id), "create")
    return smeta_to_dict(get_smeta(db, smeta.id))


@app.patch("/smetas/{smeta_id}")
def update_smeta_endpoint(
    smeta_id: int,
    payload: SmetaIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_smeta_access(db, smeta_id, user, write=True)
    raw_data = payload.model_dump()
    raw_data["tax_mode"] = raw_data.get("tax_mode") if raw_data.get("tax_mode") in {"none", "vat_added", "vat_included"} else "none"
    raw_data["section_adjustments"] = json.dumps(
        parse_section_adjustments(raw_data.get("section_adjustments", {})),
        ensure_ascii=False,
    )
    data = {key: (value if key in {"tax_rate"} else str(value).strip()) for key, value in raw_data.items()}
    data["name"] = data["name"] or "Без названия"
    smeta = update_smeta(db, smeta_id, data)
    if not smeta:
        raise HTTPException(status_code=404, detail="Смета не найдена")
    create_smeta_revision(db, get_smeta(db, smeta_id), "update")
    return smeta_to_dict(get_smeta(db, smeta_id))


@app.post("/smetas/{smeta_id}/branch")
def branch_smeta_endpoint(smeta_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    source = require_smeta_access(db, smeta_id, user)
    branch = clone_smeta(db, smeta_id, f"{source.name} - вариант")
    branch.owner_id = user.id
    db.commit()
    create_smeta_revision(db, get_smeta(db, branch.id), "branch")
    return smeta_to_dict(get_smeta(db, branch.id))


@app.delete("/smetas/{smeta_id}")
def delete_smeta_endpoint(smeta_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    smeta = require_smeta_access(db, smeta_id, user, write=True)
    if not user.is_admin and normalized_owner_id(smeta) != user.id:
        raise HTTPException(status_code=403, detail="Удалять может только владелец")
    if not delete_smeta(db, smeta_id):
        raise HTTPException(status_code=404, detail="Смета не найдена")
    return {"status": "ok"}


@app.get("/smetas/{smeta_id}")
def read_smeta(smeta_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    smeta = require_smeta_access(db, smeta_id, user)
    return smeta_to_dict(smeta)


@app.post("/smetas/{smeta_id}/share")
def share_smeta(
    smeta_id: int,
    payload: ShareIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    smeta = require_smeta_access(db, smeta_id, user, write=True)
    if not user.is_admin and normalized_owner_id(smeta) != user.id:
        raise HTTPException(status_code=403, detail="Делиться может только владелец")
    permission = payload.permission if payload.permission in {"view", "edit"} else "view"
    target_email = payload.email.strip().lower()
    target = db.query(User).filter(User.email == target_email).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    access = (
        db.query(SmetaAccess)
        .filter(SmetaAccess.smeta_id == smeta_id, SmetaAccess.user_id == target.id)
        .first()
    )
    if access:
        access.permission = permission
    else:
        db.add(SmetaAccess(smeta_id=smeta_id, user_id=target.id, permission=permission))
    db.commit()
    return {"status": "ok", "email": target.email, "permission": permission}


@app.get("/smetas/{smeta_id}/revisions")
def list_smeta_revisions_endpoint(
    smeta_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_smeta_access(db, smeta_id, user)
    revisions = get_smeta_revisions(db, smeta_id)
    return [
        {
            "id": revision.id,
            "smeta_id": revision.smeta_id,
            "label": revision.label or "",
            "created_at": revision.created_at.isoformat() if revision.created_at else None,
        }
        for revision in revisions
    ]


@app.post("/smetas/{smeta_id}/revisions/{revision_id}/restore")
def restore_smeta_revision_endpoint(
    smeta_id: int,
    revision_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_smeta_access(db, smeta_id, user, write=True)
    smeta = restore_smeta_revision(db, smeta_id, revision_id)
    if not smeta:
        raise HTTPException(status_code=404, detail="Версия не найдена")
    create_smeta_revision(db, smeta, "restore")
    return smeta_to_dict(get_smeta(db, smeta_id))


def smeta_check_issues(smeta):
    issues = []
    if not smeta:
        return issues
    for item in smeta.items:
        if (item.unit_price or 0) == 0:
            issues.append(f"Нулевая цена: {item.section or 'Раздел'} / {item.name}")
        if (item.quantity or 0) <= 0:
            issues.append(f"Некорректное количество: {item.name}")
    summary = smeta_equipment_summary(smeta)
    for kind, data in summary.items():
        matching_work = [
            item
            for item in smeta.items
            if (item.section or "") == "Монтажные работы" and work_kind(item) == kind
        ]
        work_quantity = sum(item.quantity or 0 for item in matching_work)
        if work_quantity < data["quantity"]:
            issues.append(f"Монтаж «{kind}» меньше оборудования: {work_quantity:g} из {data['quantity']:g}")
    return issues


def grouped_smeta_items(smeta):
    groups = []
    for section in DEFAULT_SECTIONS:
        items = [item for item in smeta.items if (item.section or "Оборудование") == section]
        if items:
            groups.append((section, items))
    extra_items = [
        item
        for item in smeta.items
        if (item.section or "Оборудование") not in DEFAULT_SECTIONS
    ]
    if extra_items:
        groups.append(("Прочее", extra_items))
    return groups


def org_rows(smeta):
    return [
        ("Заказчик", smeta.customer_name or "", smeta.customer_details or ""),
        ("Исполнитель", smeta.contractor_name or "", smeta.contractor_details or ""),
        ("Согласующий", smeta.approver_name or "", smeta.approver_details or ""),
    ]


def safe_filename(value):
    value = re.sub(r"[^0-9a-zA-Zа-яА-ЯёЁ._ -]+", "_", value or "smeta").strip()
    return value[:80] or "smeta"


def build_smeta_workbook(smeta):
    financials = smeta_financials(smeta)
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws.sheet_view.showGridLines = False
    details_ws = wb.create_sheet("Характеристики")
    details_ws.sheet_view.showGridLines = False

    widths = [5, 42, 56, 8, 9, 12, 13]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    def solid(color):
        return PatternFill(fill_type="solid", fgColor=f"FF{color}")

    title_fill = solid("173B57")
    section_fills = {
        "Оборудование": solid("B6D7A8"),
        "Монтажные работы": solid("F6D776"),
        "Пусконаладочные работы": solid("D9C2E9"),
        "Кабельные линии": solid("C9E2B3"),
        "Материалы и расходники": solid("F4C7B6"),
        "Доставка и логистика": solid("BDD7EE"),
        "Проектирование": solid("D9D9D9"),
        "Прочее": solid("E2E2E2"),
    }
    item_fills = {
        "Оборудование": solid("D9EAD3"),
        "Монтажные работы": solid("FFF2CC"),
        "Пусконаладочные работы": solid("EADCF8"),
        "Кабельные линии": solid("E2F0D9"),
        "Материалы и расходники": solid("FCE4D6"),
        "Доставка и логистика": solid("DDEBF7"),
        "Проектирование": solid("EDEDED"),
        "Прочее": solid("F2F2F2"),
    }
    section_fill = solid("DDEBF7")
    header_fill = solid("D9EAF7")
    org_fill = solid("EEF3F7")
    org_role_fill = solid("DDEAF6")
    total_fill = solid("1F4E78")
    total_light_fill = solid("E8F3FF")
    signature_fill = solid("FAFAFA")
    thin = Side(style="thin", color="9FB2C3")
    medium = Side(style="medium", color="6E879F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=medium, right=thin, top=medium, bottom=thin)
    money_format = '#,##0.00 "₽"'

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45

    ws["A1"] = f"Смета: {smeta.name}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["F1"] = "Итого"
    ws["G1"] = financials["total"]
    for col in range(1, 8):
        cell = ws.cell(1, col)
        cell.fill = title_fill if col <= 5 else total_light_fill
        cell.border = border
        if col <= 5:
            cell.font = Font(bold=True, size=16, color="FFFFFF")
        else:
            cell.font = Font(bold=True, size=13 if col == 6 else 15, color="173B57")
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws["G1"].number_format = money_format
    ws.row_dimensions[1].height = 28

    row = 2
    for col, value in enumerate(["Роль", "Организация", "Реквизиты", "", "", "", ""], 1):
        ws.cell(row, col, value)
    for col in range(1, 8):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, color="173B57")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    for role, name, details in org_rows(smeta):
        ws.cell(row, 1, role)
        ws.cell(row, 2, name or "Не заполнено")
        ws.cell(row, 3, details or "")
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.fill = org_role_fill if col == 1 else org_fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 1).font = Font(bold=True)
        ws.row_dimensions[row].height = 20 if not details else 34
        row += 1

    row += 1
    table_header_row = row
    headers = ["№", "Позиция", "Характеристики", "Ед.", "Кол-во", "Цена", "Сумма"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color="173B57")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24
    row += 1

    detail_rows = [["№", "Раздел", "Позиция", "Полные характеристики", "Источник"]]
    for section_index, (section, items) in enumerate(grouped_smeta_items(smeta)):
        if section_index > 0:
            row += 2
        section_total = sum(item_total(item, smeta) for item in items)
        ws.cell(row, 1, section)
        ws.cell(row, 7, section_total)
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True)
            cell.fill = section_fills.get(section, section_fill)
            cell.border = section_border if col == 1 else border
            cell.alignment = Alignment(vertical="center")
        ws.cell(row, 7).number_format = money_format
        row += 1
        section_item_no = 1
        for item in items:
            ws.cell(row, 1, section_item_no)
            ws.cell(row, 2, item.name)
            short_characteristics = clean_text(
                summarize_characteristics(item.characteristics or "", max_lines=1, line_length=135)
            )
            ws.cell(row, 3, short_characteristics)
            ws.cell(row, 4, item.unit or "ед.")
            ws.cell(row, 5, item.quantity or 0)
            ws.cell(row, 6, effective_unit_price(item, smeta))
            ws.cell(row, 7, f"=E{row}*F{row}")
            for col in range(1, 8):
                cell = ws.cell(row, col)
                cell.fill = item_fills.get(section, solid("FFFFFF"))
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row, 4).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="top")
            ws.row_dimensions[row].height = 42 if short_characteristics else 30
            ws.cell(row, 5).number_format = "0"
            ws.cell(row, 6).number_format = money_format
            ws.cell(row, 7).number_format = money_format
            detail_rows.append(
                [section_item_no, item.section or "", item.name, item.characteristics or "", item.source or ""]
            )
            section_item_no += 1
            row += 1

    last_item_row = row - 1
    row += 1
    if financials["tax_amount"] > 0 and (smeta.tax_mode or "") == "vat_added":
        ws.cell(row, 1, "Итого без НДС")
        ws.cell(row, 7, financials["subtotal"])
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="173B57")
            cell.fill = total_light_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == 6 else "center", vertical="center")
        ws.cell(row, 7).number_format = money_format
        row += 1
        ws.cell(row, 1, f"НДС {float(smeta.tax_rate or 0):g}%")
        ws.cell(row, 7, financials["tax_amount"])
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="173B57")
            cell.fill = total_light_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == 6 else "center", vertical="center")
        ws.cell(row, 7).number_format = money_format
        row += 1
    elif financials["tax_amount"] > 0 and (smeta.tax_mode or "") == "vat_included":
        ws.cell(row, 1, f"В том числе НДС {float(smeta.tax_rate or 0):g}%")
        ws.cell(row, 7, financials["tax_amount"])
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="173B57")
            cell.fill = total_light_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == 6 else "center", vertical="center")
        ws.cell(row, 7).number_format = money_format
        row += 1

    ws.cell(row, 1, "Полная сумма сметы")
    ws.cell(row, 7, financials["total"])
    for col in range(1, 8):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="right" if col == 6 else "center", vertical="center")
    ws.cell(row, 7).number_format = money_format
    ws.cell(row, 7).font = Font(bold=True, size=16, color="FFFFFF")
    ws.row_dimensions[row].height = 32

    row += 3
    ws.cell(row, 1, "Заказчик").font = Font(bold=True)
    ws.cell(row, 5, "Исполнитель").font = Font(bold=True)
    for col in range(1, 8):
        ws.cell(row, col).fill = signature_fill
    row += 2
    ws.cell(row, 1, "________________ /")
    ws.cell(row, 5, "________________ /")
    for col in range(1, 8):
        ws.cell(row, col).fill = signature_fill

    ws.freeze_panes = f"A{table_header_row + 1}"
    ws.auto_filter.ref = f"A{table_header_row}:G{last_item_row}"

    for index, values in enumerate(detail_rows, 1):
        for col, value in enumerate(values, 1):
            details_ws.cell(index, col, value)
            details_ws.cell(index, col).alignment = Alignment(wrap_text=True, vertical="top")
            details_ws.cell(index, col).border = border
        if index == 1:
            for col in range(1, 6):
                details_ws.cell(index, col).font = Font(bold=True, color="173B57")
                details_ws.cell(index, col).fill = header_fill
    for index, width in enumerate([6, 20, 42, 90, 36], 1):
        details_ws.column_dimensions[get_column_letter(index)].width = width
    details_ws.freeze_panes = "A2"

    return wb


def build_smeta_print_html(smeta):
    financials = smeta_financials(smeta)
    total = financials["total"]
    org_html = "".join(
        f"<tr><th>{escape(role)}</th><td>{escape(name)}</td><td>{escape(details).replace(chr(10), '<br>')}</td></tr>"
        for role, name, details in org_rows(smeta)
    )
    rows = []
    for section, items in grouped_smeta_items(smeta):
        section_total = sum(item_total(item, smeta) for item in items)
        rows.append(
            f"<tr class='section'><td colspan='6'>{escape(section)}</td><td>{section_total:,.2f}</td></tr>"
        )
        item_no = 1
        for item in items:
            price = effective_unit_price(item, smeta)
            total_item = item_total(item, smeta)
            rows.append(
                "<tr>"
                f"<td>{item_no}</td>"
                f"<td>{escape(item.name or '')}</td>"
                f"<td>{escape(item.characteristics or '').replace(chr(10), '<br>')}</td>"
                f"<td>{escape(item.unit or 'ед.')}</td>"
                f"<td>{item.quantity:g}</td>"
                f"<td>{price:,.2f}</td>"
                f"<td>{total_item:,.2f}</td>"
                "</tr>"
            )
            item_no += 1
    tax_html = ""
    if financials["tax_amount"] > 0 and (smeta.tax_mode or "") == "vat_added":
        tax_html = (
            f"<div class='subtotal'>Итого без НДС: {financials['subtotal']:,.2f} ₽</div>"
            f"<div class='subtotal'>НДС {float(smeta.tax_rate or 0):g}%: {financials['tax_amount']:,.2f} ₽</div>"
        )
    elif financials["tax_amount"] > 0 and (smeta.tax_mode or "") == "vat_included":
        tax_html = f"<div class='subtotal'>В том числе НДС {float(smeta.tax_rate or 0):g}%: {financials['tax_amount']:,.2f} ₽</div>"
    return f"""
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <title>Смета {escape(smeta.name)}</title>
  <style>
    body {{ font-family: Arial, sans-serif; color: #111827; margin: 28px; }}
    .actions {{ margin-bottom: 16px; }}
    button {{ background: #2364aa; color: white; border: 0; border-radius: 6px; padding: 10px 14px; }}
    h1 {{ font-size: 22px; margin: 0 0 14px; }}
    table {{ width: 100%; border-collapse: collapse; margin: 14px 0; }}
    th, td {{ border: 1px solid #cfd8e3; padding: 7px; vertical-align: top; font-size: 12px; }}
    th {{ background: #eaf2f8; }}
    .section td {{ background: #d9eaf7; font-weight: bold; }}
    .subtotal {{ text-align: right; font-size: 14px; font-weight: bold; margin-top: 8px; }}
    .total {{ text-align: right; font-size: 18px; font-weight: bold; margin-top: 12px; }}
    .sign {{ display: grid; grid-template-columns: 1fr 1fr; gap: 80px; margin-top: 42px; }}
    @media print {{ .actions {{ display: none; }} body {{ margin: 12mm; }} }}
  </style>
</head>
<body>
  <div class="actions"><button onclick="window.print()">Печать / сохранить в PDF</button></div>
  <h1>Смета: {escape(smeta.name)}</h1>
  <table>
    <thead><tr><th>Роль</th><th>Организация</th><th>Реквизиты</th></tr></thead>
    <tbody>{org_html}</tbody>
  </table>
  <table>
    <thead><tr><th>№</th><th>Позиция</th><th>Характеристики</th><th>Ед.</th><th>Кол-во</th><th>Цена</th><th>Сумма</th></tr></thead>
    <tbody>{''.join(rows)}</tbody>
  </table>
  {tax_html}
  <div class="total">Итого: {total:,.2f} ₽</div>
  <div class="sign"><div>Заказчик: __________________ /</div><div>Исполнитель: __________________ /</div></div>
</body>
</html>
"""


@app.post("/smetas/{smeta_id}/check")
def check_smeta_endpoint(smeta_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    require_smeta_access(db, smeta_id, user, write=True)
    smeta, results = validate_and_fix_smeta(db, smeta_id)
    if not smeta:
        raise HTTPException(status_code=404, detail="Смета не найдена")
    refreshed = get_smeta(db, smeta_id)
    return {
        "smeta": smeta_to_dict(refreshed),
        "results": results,
        "issues": smeta_check_issues(refreshed),
        "summary": smeta_equipment_summary(refreshed),
    }


@app.get("/smetas/{smeta_id}/export.xlsx")
def export_smeta_xlsx(smeta_id: int, token: str = Query(default=""), db: Session = Depends(get_db)):
    payload = decode_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Нужна авторизация")
    user = db.query(User).filter(User.id == int(payload.get("sub") or 0)).first()
    if not user:
        raise HTTPException(status_code=401, detail="Пользователь не найден")
    smeta = require_smeta_access(db, smeta_id, user)
    workbook = build_smeta_workbook(smeta)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = quote(f"{safe_filename(smeta.name)}.xlsx")
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.get("/smetas/{smeta_id}/print", response_class=HTMLResponse)
def print_smeta(smeta_id: int, token: str = Query(default=""), db: Session = Depends(get_db)):
    payload = decode_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Нужна авторизация")
    user = db.query(User).filter(User.id == int(payload.get("sub") or 0)).first()
    if not user:
        raise HTTPException(status_code=401, detail="Пользователь не найден")
    smeta = require_smeta_access(db, smeta_id, user)
    return HTMLResponse(build_smeta_print_html(smeta))


@app.post("/smetas/{smeta_id}/items")
def create_smeta_item_endpoint(
    smeta_id: int,
    payload: SmetaItemIn,
    material_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    smeta = require_smeta_access(db, smeta_id, user, write=True)

    item_data = payload.model_dump()
    if material_id is not None:
        material = get_material(db, material_id)
        if not material:
            raise HTTPException(status_code=404, detail="Материал не найден")
        material_type = material.item_type or classify_catalog_item(material.name, material.source)
        item_data["quantity"] = normalize_quantity(item_data.get("quantity", 1))
        item_data.update(
            {
                "item_type": material_type,
                "section": default_section_for_type(material_type),
                "name": material.name,
                "characteristics": summarize_characteristics(material.characteristics or ""),
                "unit": material.unit or "",
                "unit_price": material.price,
                "source": material.source or "",
            }
        )

    added_item = add_smeta_item(db, smeta_id, item_data)
    if is_equipment_smeta_item(added_item):
        validate_and_fix_smeta(db, smeta_id)
    create_smeta_revision(db, get_smeta(db, smeta_id), "add item")
    return smeta_to_dict(get_smeta(db, smeta_id))


@app.delete("/smetas/{smeta_id}/items/{item_id}")
def delete_smeta_item_endpoint(
    smeta_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_smeta_access(db, smeta_id, user, write=True)
    if not delete_smeta_item(db, smeta_id, item_id):
        raise HTTPException(status_code=404, detail="Позиция не найдена")
    create_smeta_revision(db, get_smeta(db, smeta_id), "delete item")
    return smeta_to_dict(get_smeta(db, smeta_id))


@app.patch("/smetas/{smeta_id}/items/{item_id}")
def update_smeta_item_endpoint(
    smeta_id: int,
    item_id: int,
    payload: SmetaItemIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_smeta_access(db, smeta_id, user, write=True)
    data = payload.model_dump()
    data["quantity"] = normalize_quantity(data["quantity"])
    item = update_smeta_item(db, smeta_id, item_id, data)
    if not item:
        raise HTTPException(status_code=404, detail="Позиция не найдена")
    if is_equipment_smeta_item(item):
        validate_and_fix_smeta(db, smeta_id)
    create_smeta_revision(db, get_smeta(db, smeta_id), "update item")
    return smeta_to_dict(get_smeta(db, smeta_id))


@app.get("/settings/ai")
def get_ai_settings(user: User = Depends(get_current_user)):
    require_admin_user(user)
    return public_settings(read_settings())


@app.post("/settings/ai")
def save_ai_settings(payload: AiSettingsIn, user: User = Depends(get_current_user)):
    require_admin_user(user)
    current = read_settings()
    settings = {
        "base_url": payload.base_url.strip().rstrip("/") or "https://api.vsegpt.ru/v1",
        "api_key": payload.api_key.strip() or current.get("api_key", ""),
        "model": payload.model.strip(),
        "assistant_prompt": payload.assistant_prompt.strip(),
    }
    write_settings(settings)
    return public_settings(settings)


@app.get("/settings/ai/models")
def get_ai_models(user: User = Depends(get_current_user)):
    require_admin_user(user)
    settings = read_settings()
    try:
        with httpx.Client(timeout=30) as client:
            response = client.get(endpoint(settings["base_url"], "models"), headers=provider_headers(settings))
            response.raise_for_status()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=http_error_detail(exc, "Не удалось получить список моделей")) from exc

    data = response.json()
    models = data.get("data", data if isinstance(data, list) else [])
    return {"models": [normalize_model(model) for model in models if isinstance(model, dict)]}


@app.post("/ai/recommend")
def ai_recommend(prompt: str = Query(...)):
    settings = read_settings()
    if settings.get("api_key") and settings.get("model"):
        payload = {
            "model": settings["model"],
            "messages": [
                {
                    "role": "system",
                    "content": "Ты помощник сметчика. Отвечай кратко, практично и по-русски.",
                },
                {"role": "user", "content": prompt},
            ],
        }
        try:
            with httpx.Client(timeout=60) as client:
                response = client.post(
                    endpoint(settings["base_url"], "chat/completions"),
                    headers=provider_headers(settings),
                    json=payload,
                )
                response.raise_for_status()
            recommendation = response.json()["choices"][0]["message"]["content"]
            append_ai_audit(
                "ai_recommend",
                {
                    "prompt": prompt,
                    "provider": settings.get("base_url"),
                    "model": settings.get("model"),
                    "recommendation": recommendation,
                    "mode": "provider",
                },
            )
            return {"recommendation": recommendation}
        except httpx.HTTPError as exc:
            append_ai_audit(
                "ai_recommend",
                {
                    "prompt": prompt,
                    "provider": settings.get("base_url"),
                    "model": settings.get("model"),
                    "error": str(exc),
                    "mode": "provider_error",
                },
            )
            raise HTTPException(status_code=502, detail=http_error_detail(exc, "AI-провайдер отклонил запрос")) from exc
        except (KeyError, IndexError) as exc:
            append_ai_audit(
                "ai_recommend",
                {
                    "prompt": prompt,
                    "provider": settings.get("base_url"),
                    "model": settings.get("model"),
                    "error": str(exc),
                    "mode": "provider_bad_payload",
                },
            )
            raise HTTPException(status_code=502, detail=f"AI-провайдер не ответил корректно: {exc}") from exc
    recommendation = simple_ai_assistant(prompt)
    append_ai_audit(
        "ai_recommend",
        {
            "prompt": prompt,
            "provider": "fallback",
            "model": None,
            "recommendation": recommendation,
            "mode": "local_fallback",
        },
    )
    return {"recommendation": recommendation}


@app.post("/ai/command")
def ai_command(payload: AiCommandIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    selected = get_smeta(db, payload.smeta_id) if payload.smeta_id else None
    new_smeta_request = looks_like_new_smeta_request(payload.prompt)
    if new_smeta_request:
        inferred_name = infer_smeta_name(payload.prompt, "", user)
        name = resolve_ai_smeta_name(db, payload.prompt, inferred_name, user)
        smeta = create_smeta(db, name, {"owner_id": user.id} if user else None)
        build_results = auto_build_project_smeta(db, smeta.id, payload.prompt, user)
        create_smeta_revision(db, get_smeta(db, smeta.id), "create")
        log_ai_command(
            "auto_create",
            user,
            payload.prompt,
            payload=payload.model_dump(),
            reply=f"Создал смету «{name}».",
            results=[f"Создал смету «{name}»", *build_results],
            selected_smeta_id=smeta.id,
            extra={"created_name": name},
        )
        return {
            "reply": f"Создал смету «{name}».",
            "results": [f"Создал смету «{name}»", *build_results],
            "selected_smeta_id": smeta.id,
            "smetas": [smeta_to_dict(smeta) for smeta in get_visible_smetas(db, user)],
        }
    if payload.smeta_id:
        require_smeta_access(db, payload.smeta_id, user)
    if payload.smeta_id and looks_like_extend_smeta_request(payload.prompt):
        require_smeta_access(db, payload.smeta_id, user, write=True)
        smeta = get_smeta(db, payload.smeta_id)
        if not smeta:
            raise HTTPException(status_code=404, detail="Смета не найдена")
        results = auto_build_project_smeta(db, payload.smeta_id, payload.prompt, user)
        create_smeta_revision(db, get_smeta(db, payload.smeta_id), "add item")
        log_ai_command(
            "auto_extend",
            user,
            payload.prompt,
            payload=payload.model_dump(),
            reply="Дополнил смету по вашему запросу.",
            results=results,
            selected_smeta_id=payload.smeta_id,
            extra={"base_smeta_id": payload.smeta_id},
        )
        return {
            "reply": "Дополнил смету по вашему запросу.",
            "results": results,
            "selected_smeta_id": payload.smeta_id,
            "smetas": [smeta_to_dict(smeta) for smeta in get_visible_smetas(db, user)],
        }
    count_answer = answer_count_question(payload.prompt, selected)
    if count_answer:
        log_ai_command(
            "count_question",
            user,
            payload.prompt,
            payload=payload.model_dump(),
            reply=count_answer,
            results=[],
            selected_smeta_id=payload.smeta_id,
        )
        return {
            "reply": count_answer,
            "results": [],
            "selected_smeta_id": payload.smeta_id,
            "smetas": [smeta_to_dict(smeta) for smeta in get_visible_smetas(db, user)],
        }

    if payload.smeta_id and should_validate_smeta(payload.prompt):
        require_smeta_access(db, payload.smeta_id, user, write=True)
        smeta, results = validate_and_fix_smeta(db, payload.smeta_id)
        log_ai_command(
            "validate_smeta",
            user,
            payload.prompt,
            payload=payload.model_dump(),
            reply="Проверил смету по всей структуре: оборудование, монтажные работы и цены из базы.",
            results=results,
            selected_smeta_id=payload.smeta_id,
        )
        return {
            "reply": "Проверил смету по всей структуре: оборудование, монтажные работы и цены из базы.",
            "results": results,
            "selected_smeta_id": payload.smeta_id,
            "smetas": [smeta_to_dict(smeta) for smeta in get_visible_smetas(db, user)],
        }

    if payload.smeta_id and should_auto_add_installation(payload.prompt):
        require_smeta_access(db, payload.smeta_id, user, write=True)
        smeta, results = add_installation_works_for_smeta(db, payload.smeta_id)
        log_ai_command(
            "auto_installation",
            user,
            payload.prompt,
            payload=payload.model_dump(),
            reply="Проверил оборудование в смете и добавил монтажные работы из базы. Если цена не найдена, поставил 0.",
            results=results,
            selected_smeta_id=payload.smeta_id,
        )
        return {
            "reply": "Проверил оборудование в смете и добавил монтажные работы из базы. Если цена не найдена, поставил 0.",
            "results": results,
            "selected_smeta_id": payload.smeta_id,
            "smetas": [smeta_to_dict(smeta) for smeta in get_visible_smetas(db, user)],
        }

    settings = read_settings()
    system_prompt = "\n\n".join(
        [
            merged_ai_system_prompt(settings),
            (
                "Ты управляешь сметным приложением. Верни только JSON-объект без markdown: "
                '{"reply":"короткий ответ пользователю","actions":[...]} . '
                "Разрешенные actions: "
                "create_smeta {name}; delete_smeta {smeta_id}; "
                "add_item {smeta_id, section, name, characteristics, unit, quantity, unit_price, source}; "
                "update_item {smeta_id, item_id, section, name, characteristics, unit, quantity, unit_price, source}; "
                "delete_item {smeta_id, item_id}. "
                f"Допустимые разделы: {', '.join(DEFAULT_SECTIONS)}. "
                "У тебя есть доступ к контексту smetas и work_price_examples в JSON ниже. "
                "Если пользователь просит взять цены из базы, используй work_price_examples. "
                "Если не нашел цену работы, добавь позицию с unit_price 0. "
                "Для вопросов количества используй selected_smeta_equipment_summary. "
                "Синонимы обязательны: камера=видеокамера, регистратор=видеорегистратор/NVR/DVR. "
                "Считай оборудование по всей выбранной смете, а не только по одному разделу. "
                "Если пользователь просит удалить, меняй или удаляй только явно указанную смету или позицию. "
                "Не придумывай цены, если пользователь их не дал."
            ),
        ]
    )
    equipment_names = [
        item.name
        for item in (selected.items if selected else [])
        if (item.section or "") == "Оборудование" or (item.item_type or "") == "equipment"
    ]
    work_examples = []
    for name in equipment_names[:20]:
        work, kind = find_work_price(db, name)
        work_examples.append(
            {
                "equipment": name,
                "kind": kind,
                "work": work.name if work else None,
                "unit": work.unit if work else "шт",
                "price": work.price if work else 0,
                "source": work.source if work else "not_found",
            }
        )
    context = {
        "selected_smeta_id": payload.smeta_id,
        "smetas": [
            {
                "id": smeta.id,
                "name": smeta.name,
                "total": smeta_to_dict(smeta)["total"],
                "items": [
                    {
                        "id": item.id,
                        "item_type": item.item_type,
                        "section": item.section or "Оборудование",
                        "name": item.name,
                        "quantity": item.quantity,
                        "unit": item.unit or "",
                        "unit_price": item.unit_price,
                    }
                    for item in smeta.items
                ],
            }
            for smeta in get_visible_smetas(db, user)[:20]
        ],
        "selected_smeta_equipment_summary": smeta_equipment_summary(selected) if selected else {},
        "synonyms": {
            "камера": ["камера", "видеокамера"],
            "регистратор": ["регистратор", "видеорегистратор", "NVR", "DVR"],
        },
        "work_price_examples": work_examples,
        "user_request": payload.prompt,
    }
    decision = call_ai_object(system_prompt, json.dumps(context, ensure_ascii=False))
    actions = decision.get("actions", [])
    if not isinstance(actions, list):
        actions = []
    if not actions and should_create_smeta(payload.prompt):
        actions = [{
            "action": "create_smeta",
            "name": infer_smeta_name(payload.prompt, decision.get("reply") or "", user),
        }]
    selected_smeta_id, results = execute_ai_actions(db, actions, payload.smeta_id, user, payload.prompt)
    log_ai_command(
        "llm_actions",
        user,
        payload.prompt,
        payload=payload.model_dump(),
        reply=decision.get("reply") or "Готово.",
        results=results,
        selected_smeta_id=selected_smeta_id,
        extra={
            "decision": decision,
            "actions": actions,
            "used_fallback_create": bool(not actions and should_create_smeta(payload.prompt)),
        },
    )
    return {
        "reply": decision.get("reply") or "Готово.",
        "results": results,
        "selected_smeta_id": selected_smeta_id,
        "smetas": [smeta_to_dict(smeta) for smeta in get_visible_smetas(db, user)],
    }
